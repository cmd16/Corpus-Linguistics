"""
Abstract Nouns
Written by Catherine DeJager

Get the abstract nouns in a corpus by searching for various suffixes as found in the website below.
    https://learningenglishgrammar.wordpress.com/suffixes/suffixes-and-how-they-form-abstract-nouns/
Dependencies: openpyxl (if you want to output results to spreadsheet)
"""

# imports
import openpyxl
import os

# global variables
test = True
suffix_list = ['ance', 'ances', 'cy', 'cies', 'dom', 'doms', 'ence', 'ences', 'ness',
               'nesses', 'esses', 'hood', 'hoods', 'ice', 'ices', 'ion', 'ions', 'ism',
               'isms', 'ity', 'itys', 'ment', 'ments', 'ty', 'ties', 'tude', 'tudes', 'ure', 'ures']


# function definitions
def get_abstract_nouns_from_wordlist(infile):
    """
    Returns a dictionary of abstract nouns, with a separate entry for each suffix.
    :param infile: a .txt file containing a word list created from AntConc
    :return: a dictionary of abstract nouns, with a separate entry for each suffix. Also the number of tokens.
    """
    abstract_noun_dict = {suffix: [] for suffix in suffix_list}  # each suffix has a blank array (for word and frequency)
    types = 0
    tokens = 0
    linenum = 0
    for line in infile:
        if linenum < 3:  # first 3 lines contain metadata
            if linenum == 0:
                types = int(line[line.index("types")+7:])
            elif linenum == 1:
                tokens = int(line[line.index("tokens")+8:])
            linenum += 1
            continue
        data = str(line).split()
        word = data[2]
        frequency = int(data[1])
        norm_frequency = frequency * 1000000 / tokens
        for idx in range(len(suffix_list)):
            suffix = suffix_list[idx]
            if word.endswith(suffix):
                abstract_noun_dict[suffix].append([word, frequency, norm_frequency])
                break
        linenum += 1
    infile.close()
    return [abstract_noun_dict, types, tokens]


def get_abstract_nouns_from_txt(infile):
    """
    Returns a dictionary of abstract nouns, with a separate entry for each suffix.
    Note: this seems to work but doesn't pass my current tests, so use at your own risk.
    :param infile: a .txt file with words in it.
    :return: a dictionary of abstract nouns, with a separate entry for each suffix.
    """
    """Returns a dictionary of abstract nouns, with a separate entry for each suffix.
    Preconditions: infile refers to a .txt file with words in it."""
    # TODO: CHANGE THIS LATER TO USE MULTIDIMENSIONAL ARRAY FOR WORDS

    abstract_noun_dict = {suffix: {} for suffix in suffix_list}  # each suffix has a blank dictionary (for word and frequency)
    for line in infile:
        words = str(line).split()
        for word in words:
            for suffix in suffix_list:
                if word.endswith(suffix):
                    if word in abstract_noun_dict[suffix]:
                        abstract_noun_dict[suffix][word] += 1
                    else:
                        abstract_noun_dict[suffix][word] = 1
                    break
    infile.close()
    return abstract_noun_dict


def store_spreadsheet(aDict, filename="Abstract Nouns Spreadsheet"):
    """
    Stores the abstract nouns and their frequencies in a spreadsheet.
    :param aDict: a dictionary created by running get_abstract_nouns
    :param filename: the name of the spreadsheet to store the results in
    :return:
    """
    wb = openpyxl.Workbook()
    ws = wb.active  # get a handle to the sheet in the workbook

    ws['A1'] = "Abstract Nouns Spreadsheet"
    for col in range(1, len(suffix_list) * 2 + 1, 2):  # go up by two because the suffixes need to be separated. add one because indexing starts at 1 in openpyxl
        suffix = suffix_list[(col - 1) // 2]
        current_suffix_entry = aDict[suffix]
        ws.cell(row=2, column=col).value = "-" + suffix
        for row in range(3, len(current_suffix_entry) + 3):  # this number is equal to the number of unique words with that suffix
            idx = row - 3
            word = current_suffix_entry[idx][0]
            ws.cell(row=row, column=col).value = word
            ws.cell(row=row, column=col+1).value = current_suffix_entry[idx][1]  # store the frequency
    if not filename.endswith(".xlsx"):
        wb.save(filename + ".xlsx")  # add the type extension if not included
    else:
        wb.save(filename)


def sort_abstract_nouns(aDict, sort_key='frequencyhi'):
    """
    Sort (in place) a dictionary that was constructed using get_abstract_nouns
    :param aDict: a dictionary constructed using get_abstract_nouns
    :param sort_key: a string representing which way to sort the words/frequencies
    :return:
    """
    if sort_key == "frequencyhi":
        for suffix in suffix_list:
           aDict[suffix].sort(key=lambda x: x[1], reverse=True)
    elif sort_key == "frequencylo":
        for suffix in suffix_list:
            aDict[suffix].sort(key=lambda x: x[1])
    elif sort_key == "alphalo":
        for suffix in suffix_list:
           aDict[suffix].sort()
    elif sort_key == "alphahi":
        for suffix in suffix_list:
           aDict[suffix].sort(reverse=True)
    elif sort_key == "alphawordendlo":
        raise NotImplementedError
    elif sort_key == "alphawordendhi":
        raise NotImplementedError
    else:
        print('Sorting type invalid. Try again with "frequencyhi" (high to low), "frequencylo" (low to high), "alphahi", "alphalo", '
              '"alphawordendlo", or "alphawordendhi"')


def abstract_nouns_store_count(Dict_and_data, outfilename, num=10):
    """
    Count the most frequent num words in each
    :param Dict_and_data:  a suffix Dict and type and token counts as created by get_abstract_nouns
    :param outfilename: the txt file to store the abstract noun info in
    :param num: the number of abstract nouns to store from each category
    :return:
    :prerequisite: the Dict in Dict_and_data is sorted by frequencyhi (high to low)
    """
    aDict, wordlist_types, wordlist_tokens = Dict_and_data
    total_types = 0
    total_tokens = 0
    count_dict = {suffix: [] for suffix in suffix_list}  # suffix: [types, tokens, [top_num]]
    for suffix in suffix_list:
        entry = aDict[suffix]
        types = len(entry)
        total_types += types
        count_dict[suffix].append(types)
        tokens = 0
        for freq_tup in entry:
            tokens += freq_tup[1]
        total_tokens += tokens
        count_dict[suffix].append(tokens)
        count_dict[suffix].append(entry[0:num+1])
    try:
        norm_types = total_types*1000000/wordlist_types
        norm_tokens = total_tokens*1000000/wordlist_tokens
    except ZeroDivisionError:
        norm_types = 0
        norm_tokens = 0
    f_out = open(outfilename, "w")
    f_out.write("# Word types: %d\t%f\n" % (total_types, norm_types))
    f_out.write("# Word tokens: %d\t%f\n" % (total_tokens, norm_tokens))
    for suffix in suffix_list:
        count_entry = count_dict[suffix]
        f_out.write("# %s types: %d\n" % (suffix, count_entry[0]))
        f_out.write("# %s tokens: %d\n" % (suffix, count_entry[1]))
        for freq_tup in count_entry[2]:
            f_out.write("%s:\t%d\t%f\n" % (freq_tup[0], freq_tup[1], freq_tup[2]))
    f_out.close()


def walk_dir_abstract_nouns_spreadsheet(path, keyword="_antconc"):
    """
    Walks through a directory and makes a spreadsheet of abstract nouns for each file in the directory.
    Preconditions: _antconc is in the filename of each wordlist file
    :param path: the path to a directory containing- AntConc wordlists and/or folders containing AntConc wordlists
    :param keyword: the word that indicates the file is a wordlist
    :return:
    """
    for item in os.walk(path):
        print(item)
        os.chdir(item[0])  # change to the directory you are looking at. useful for reading and writing files
        for filename in item[2]:
            if filename.endswith(".txt") and keyword in filename:
                this_dict = get_abstract_nouns_from_wordlist(open(filename))[0]  # don't care about types and tokens
                sort_abstract_nouns(this_dict, "frequencyhi")
                store_spreadsheet(this_dict, filename[:filename.index(".txt")] + "_abstract_nouns")


def walk_dir_count_abstract_nouns(path, out_dir, keyword="_python"):
    """
    Walks through a directory of wordlists and counts the abstract nouns, then stores this information in a file for each wordlist
    in the directory specified
    :param path: the directory containing wordlists
    :param out_dir: the directory to store the results in
    :param keyword: the keyword that indicates the file is a wordlist
    :return:
    """
    # print(os.listdir(path))
    for item in os.walk(path):
        # print("item", item)
        os.chdir(item[0])  # change to the directory you are looking at. useful for reading and writing files
        for filename in item[2]:
            if filename.endswith(".txt"):
                print(filename)
                dict_and_data = get_abstract_nouns_from_wordlist(open(filename))
                sort_abstract_nouns(dict_and_data[0], "frequencyhi")
                abstract_nouns_store_count(dict_and_data, os.path.join(out_dir, filename[:filename.index(keyword)] + "_abstract nouns_python.txt"))


