import csv
import os
import math
import openpyxl
import operator
import re
import numpy as np
import matplotlib.pyplot as plt
import time

stats = ['work_id', 'title', 'rating', 'category', 'fandom', 'relationship', 'character', 'additional tags', 'language',
         'published', 'status', 'status date', 'words', 'chapters', 'comments', 'kudos', 'bookmarks', 'hits', 'body']

csv.field_size_limit(1000000000)  # sys.maxsize


def ao3_txt_from_csv(csv_in, path, lang="English", restart=0, end=0):
    """
    Gets the text of each fanfic from a csv file created using ao3_get_fanfics.py
    :param csv_in: the name of a csv file created using ao3_get_fanfics.py
    :param path: path to where you want to store the txt files
    :param restart: the id to restart at
    :return:
    """
    f_in = open(csv_in, 'r+')
    reader = csv.reader(f_in)
    try:
        os.chdir(path)
    except FileNotFoundError:
        os.mkdir(path)
        os.chdir(path)
    next(reader, None)  # skip the headers
    if restart:
        found = False
    else:
        found = True
    for row in reader:
        work_id = row[0]  # stats.index("work_id")
        """if end and int(work_id) == end:
            break"""
        if restart and not found:
            id_ = int(work_id)
            if id_ == restart:
                found = True
                print("found", id_)
            else:
                print("skipping already processed fic")
        if found:
            print("ao3_txt_from_csv processing", work_id)
            if row[8] == lang:
                # print(work_id)
                fandom = row[4]
                # print(fandom)
                filename = work_id + ".txt"
                if os.path.isfile(filename):
                    outfile = open(filename, "w")
                    outfile.write(row[-1])  # stats.index("body")
                    outfile.close()
            else:
                print(work_id, row[8])
    f_in.close()


def ao3_count_stats_from_csv(csv_in):
    """
    Counts the statistics from a csv file created using ao3_get_fanfics.py
    :param csv_in: the name of a csv file created using ao3_get_fanfics.py
    :return: a dictionary containing the instances of each statistic and their frequencies
    """
    f_in = open(csv_in, 'r+')
    reader = csv.reader(f_in)
    header = next(reader)  # change if Sherlock
    # print(header)
    stat_names = ['work_id', 'title', 'rating', 'category', 'fandom', 'relationship', 'character', 'additional tags', 'language', 'published', 'status', 'status date', 'words', 'chapters', 'comments', 'kudos', 'bookmarks', 'hits', 'body']
    # header = {stat: stat_names.index(stat) for stat in stat_names}
    header = {stat: header.index(stat) for stat in header}
    for stat in ["work_id", "language", "body"]:  # TODO: add comments later
        del header[stat]  # remove the stats we don't care about
    stats = {stat: {} for stat in header}
    for row in reader:
        # print(row)
        for stat in stats:
            row_stat = row[header[stat]]
            if row_stat == 'null' or row_stat == '':
                row_stat = 0
            if stat in ["words", "kudos", "bookmarks", "comments", "hits"]:  # TODO: put chapters back later
                row_stat = int(row_stat)
                if stat == "words":
                    row_stat = (math.ceil(int(row_stat) / 100.0)) * 100
                elif stat == "hits":
                    row_stat = (math.ceil(int(row_stat) / 10.0)) * 10
            current_entry = stats[stat]
            if row_stat in current_entry:
                current_entry[row_stat] += 1
            else:
                current_entry[row_stat] = 1
    f_in.close()
    return stats


def ao3_edit_csv(csv_in, csv_out):
    """
    Create a new csv, removing any fics with that have 0 words, can't be opened (because the file doesn't exist), or are duplicates
    :param csv_in: a csv created using ao3_get_fanfics.py
    :param csv_out: the modified csv file
    :return:
    """
    f_in = open(csv_in, 'r+')
    f_out = open(csv_out, 'w')
    reader = csv.reader(f_in)
    writer = csv.writer(f_out)
    # header = next(reader)
    # writer.writerow(header)
    stat_names = ['work_id', 'title', 'rating', 'category', 'fandom', 'relationship', 'character', 'additional tags',
                  'language',
                  'published', 'status', 'status date', 'words', 'chapters', 'comments', 'kudos', 'bookmarks', 'hits']  # no column for body
    writer.writerow(stat_names)
    header = {stat: stat_names.index(stat) for stat in stat_names}
    # print(header)
    # header = {stat: header.index(stat) for stat in header}
    fanfic_dir = "/Volumes/2TB/Final_project/Fanfic_all"
    ids_seen = []
    for row in reader:
        idname = row[header['work_id']]
        if idname == "work_id":
            continue
        word_count = row[header["words"]]
        if word_count != 'null' and word_count != '' and word_count != 0 and word_count != "0":
            if idname in ids_seen:
                print("repeat", csv_in, idname)
                continue
            try:
                txt_in = open(os.path.join(fanfic_dir, str(row[0]) + ".txt"))
                for i in range(20):
                    if re.search('[a-zA-Z]', txt_in.readline()):  # search the first 20 lines for letters
                        writer.writerow(row[:len(stat_names)])  # don't need to write the body of the fic
                        ids_seen.append(idname)
                        break
                else:  # couldn't find letters
                    print(csv_in, row[0], end=", ")
                txt_in.close()
            except FileNotFoundError:
                print(csv_in, row[0], end="! ")
        else:
            print(csv_in, row[0], end="* ")
    print()


def remove_ids(csv_in, csv_out, idlist):
    """
    Take specific work ids out of a csv and store the resulting csv in a new file
    :param csv_in: a csv created using ao3_get_fanfics.py
    :param csv_out: a csv with the specified ids removed
    :param idlist: a list of ids to remove (not including ".txt")
    :return:
    """
    f_in = open(csv_in, 'r+')
    f_out = open(csv_out, 'w')
    reader = csv.reader(f_in)
    writer = csv.writer(f_out)
    header = next(reader)
    writer.writerow(header)
    stat_names = ['work_id', 'title', 'rating', 'category', 'fandom', 'relationship', 'character', 'additional tags',
                  'language',
                  'published', 'status', 'status date', 'words', 'chapters', 'comments', 'kudos', 'bookmarks', 'hits']  # no column for body
    writer.writerow(stat_names)
    header = {stat: stat_names.index(stat) for stat in stat_names}
    # print(header)
    # header = {stat: header.index(stat) for stat in header}
    fanfic_dir = "/Volumes/2TB/Final_project/Fanfic_all"
    for row in reader:
        id = row[header["work_id"]]
        if int(id) not in idlist:
            try:
                txt_in = open(os.path.join(fanfic_dir, str(row[0]) + ".txt"))
                for i in range(20):
                    if re.search('[a-zA-Z]', txt_in.readline()):  # search the first 20 lines for letters
                        writer.writerow(row[:len(row) - 1])  # don't need to write the body of the fic
                        break
                else:  # couldn't find letters
                    print(csv_in, row[0], end=", ")
                txt_in.close()
            except FileNotFoundError:
                print(csv_in, row[0], end="! ")
        else:
            print(csv_in, row[0], end="* ")
    print()


def get_numwords_ids(csv_in, out_txt_name, min_words, max_words):
    """
    Get fics with the specified number of words.
    :param csv_in: a csv created using ao3_get_fanfics.py
    :param out_txt_name: a txt file containing each id on a new line
    :param min_words: inclusive
    :param max_words: inclusive
    :return:
    """
    f_in = open(csv_in, 'r+')
    f_out = open(out_txt_name, 'a+')
    reader = csv.reader(f_in)
    header = next(reader)
    # print(header)
    header = {stat: header.index(stat) for stat in header}
    seen = []
    for row in reader:
        work_id = row[header["work_id"]]
        try:
            word_count = int(row[header["words"]])
        except ValueError:
            continue
        if word_count == 'null' or word_count == '':
            continue
        if min_words <= word_count <= max_words:
            f_out.write(work_id)
            f_out.write("\n")
            seen.append(work_id)


def get_rating_ids(csv_in, out_txt_name, rating):
    """
    Get fics with the specified rating
    :param csv_in: a csv created using ao3_get_fanfics.py
    :param out_txt_name: a txt file containing each id on a new line
    :param rating: the rating to search for
    :return:
    """
    f_in = open(csv_in, 'r+')
    f_out = open(out_txt_name, 'w')
    reader = csv.reader(f_in)
    header = next(reader)
    # print(header)
    stat_names = ['work_id', 'title', 'rating', 'category', 'fandom', 'relationship', 'character', 'additional tags',
                  'language',
                  'published', 'status', 'status date', 'words', 'chapters', 'comments', 'kudos', 'bookmarks', 'hits',
                  'body']
    header = {stat: stat_names.index(stat) for stat in stat_names}
    seen = []
    for row in reader:
        rating_str = row[header['rating']]
        if rating_str == 'null' or rating_str == '':
            continue
        if rating_str == rating:
            work_id = row[header['work_id']]
            if work_id in seen:
                print(out_txt_name, work_id)
            else:
                f_out.write(work_id)
                f_out.write("\n")
                seen.append(work_id)
    f_in.close()
    f_out.close()


def get_fandom_ids(csv_in, out_txt_name, fandom_list):
    """
    Get all the ids of fics including one or more fandoms in the fandom list, but no fandoms not in the list
    :param csv_in: a csv created using ao3_get_fanfics.py
    :param out_txt_name: a txt file containing each id on a new line
    :param fandom_list: a list of fandoms
    :return:
    """
    f_in = open(csv_in, 'r+')
    f_out = open(out_txt_name, 'w')
    reader = csv.reader(f_in)
    header = next(reader)
    # print(header)
    stat_names = ['work_id', 'title', 'rating', 'category', 'fandom', 'relationship', 'character', 'additional tags',
                  'language',
                  'published', 'status', 'status date', 'words', 'chapters', 'comments', 'kudos', 'bookmarks', 'hits',
                  'body']
    header = {stat: stat_names.index(stat) for stat in stat_names}
    seen = []
    for row in reader:
        work_id = row[header['work_id']]
        if work_id in seen:
            print(out_txt_name, work_id)
            continue
        fandom_str = row[header['fandom']]
        if fandom_str == 'null' or fandom_str == '':
            continue
        fandoms = fandom_str.split(", ")
        for fandom in fandoms:
            if fandom not in fandom_list:
                break
        else:  # if all fandoms were in the valid list
            f_out.write(row[header['work_id']])
            f_out.write("\n")
            seen.append(work_id)
    f_in.close()
    f_out.close()


def get_csv_ids(csv_in, out_txt_name):
    """
    Get all the ids from a csv
    :param csv_in: a csv created using ao3_get_fanfics.py
    :param out_txt_name: a txt file containing each id on a new line
    :return:
    """
    f_in = open(csv_in, 'r+')
    f_out = open(out_txt_name, 'w')
    reader = csv.reader(f_in)
    header = next(reader)
    # print(header)
    stat_names = ['work_id', 'title', 'rating', 'category', 'fandom', 'relationship', 'character', 'additional tags',
                  'language',
                  'published', 'status', 'status date', 'words', 'chapters', 'comments', 'kudos', 'bookmarks', 'hits',
                  'body']
    header = {stat: stat_names.index(stat) for stat in stat_names}
    seen = []
    for row in reader:
        word_str = row[header['words']]
        if word_str == 'null' or word_str == '' or word_str == 0:
            continue
        work_id = row[header["work_id"]]
        if work_id in seen:
            print(out_txt_name, work_id)
        else:
            f_out.write(work_id)
            f_out.write("\n")
    f_in.close()


def get_fandom_group_ids(csv_in, out_txt_name, fandom_groups, num_to_match=1):  # TODO: fix this so it works
    """
    Get fics with at least one fandom from each group and no fandoms that aren't in any group
    :param csv_in: a csv created using ao3_get_fanfics.py
    :param out_txt_name: a txt file containing each id on a new line
    :param fandom_groups: a list of groups of lists of fandoms (e.g., [["Harry Potter", "HP"], ["Merlin (TV)", "BBC Merlin"], ["Supernatural"]]
                            assumption: the groups are disjoint.
    :param num_to_match: how many fandoms from each group must be matched
    :return:
    """
    f_in = open(csv_in, 'r+')
    f_out = open(out_txt_name, 'w')
    reader = csv.reader(f_in)
    header = next(reader)
    # print(header)
    stat_names = ['work_id', 'title', 'rating', 'category', 'fandom', 'relationship', 'character', 'additional tags',
                  'language',
                  'published', 'status', 'status date', 'words', 'chapters', 'comments', 'kudos', 'bookmarks', 'hits',
                  'body']
    header = {stat: stat_names.index(stat) for stat in stat_names}
    for row in reader:
        fandom_str = row[header['fandom']]
        if fandom_str == 'null' or fandom_str == '':
            continue
        fandoms = fandom_str.split(", ")
        group_counts = [[x, 0] for x in fandom_groups]  # create a multidimensional array to count the number of tags in each category
        other_fandoms = False  # set a flag for if other fandoms are found
        for fandom in fandoms:
            found = False  # track if the fandom has been found in any of the groups
            for i in range(len(group_counts)):
                group, freq = group_counts[i]
                if fandom in group:
                    freq += 1
                    found = True
            if not found:
                other_fandoms = True
                break
        if not other_fandoms:
            for i in range(len(group_counts)):
                if group_counts[i][1] < num_to_match:  # if not enough fandoms in each group have been found
                    break
            else:
                f_out.write(row[header['work_id']])
                f_out.write("\n")
    f_in.close()
    f_out.close()


def get_category_ids(csv_in, out_txt_name, categories):
    """
    Get all the ids of fics with all of the relationship categories specified
    :param csv_in: a csv created using ao3_get_fanfics.py
    :param out_txt_name: a txt file containing each id on a new line
    :param categories: categories to match
    :return:
    """
    seen = []
    f_in = open(csv_in, 'r+')
    f_out = open(out_txt_name, 'w')
    reader = csv.reader(f_in)
    header = next(reader)
    # print(header)
    stat_names = ['work_id', 'title', 'rating', 'category', 'fandom', 'relationship', 'character', 'additional tags',
                  'language',
                  'published', 'status', 'status date', 'words', 'chapters', 'comments', 'kudos', 'bookmarks', 'hits',
                  'body']
    header = {stat: stat_names.index(stat) for stat in stat_names}
    for row in reader:
        category_str = row[header['category']]
        if category_str == 'null' or category_str == '':
            continue
        category_list = category_str.split(", ")
        categories.sort()
        category_list.sort()
        if categories == category_list:  # if the lists of categories are identical
            work_id = row[header['work_id']]
            if work_id in seen:
                print(out_txt_name, work_id)
            else:
                f_out.write(work_id)
                f_out.write("\n")
                seen.append(work_id)
    f_in.close()
    f_out.close()


def get_tag_ids(csv_in, out_txt_name, tags, num=1, exact=False, solo=True):  # TODO: change for else
    """
    Get all the ids of fics that match at least (num) of the tags specified
    :param csv_in: a csv created using ao3_get_fanfics.py
    :param out_txt_name: a txt file containing each id on a new line
    :param num: the number of tags that must be matched
    :param tags: the tags to match
    :param exact: if True, the tag must match exactly. Otherwise, just search for the specified tags
    :param solo: if True, no other tags are allowed
    :return:
    """
    f_in = open(csv_in, 'r+')
    f_out = open(out_txt_name, 'w')
    reader = csv.reader(f_in)
    header = next(reader)
    # print(header)
    stat_names = ['work_id', 'title', 'rating', 'category', 'fandom', 'relationship', 'character', 'additional tags',
                  'language',
                  'published', 'status', 'status date', 'words', 'chapters', 'comments', 'kudos', 'bookmarks', 'hits',
                  'body']
    header = {stat: stat_names.index(stat) for stat in stat_names}
    seen = []
    for row in reader:
        tag_str = row[header['additional tags']]
        if tag_str == 'null' or tag_str == '':
            continue
        work_id = row[header["work_id"]]
        if work_id in seen:
            print(out_txt_name, work_id)
            continue
        num_found = 0
        if exact:
            tag_list = tag_str.split(", ")
            if not solo:
                for tag in tag_list:
                    if tag not in tags:
                        break
                    else:
                        num_found += 1
                else:
                    if num_found >= num:  # if no invalid tags present and enough valid tags found
                        f_out.write(row[header['work_id']])
                        f_out.write("\n")
                        seen.append(work_id)
            else:
                for tag in tags:
                    if tag in tags:
                        num_found += 1  # TODO: improve the fact that we keep looking when we don't need to
                if num_found >= num:  # if enough valid tags found
                    f_out.write(row[header['work_id']])
                    f_out.write("\n")
                    seen.append(work_id)
        else:
            if solo:
                for tag in tags:
                    if tag in tag_str:
                        num_found += 1
                if num_found >= num:  # if enough of the tags are in the string
                    f_out.write(row[header['work_id']])
                    f_out.write("\n")
                    seen.append(work_id)
            else:
                tag_list = tag_str.split(", ")
                for tag in tags:
                    if tag in tag_str:
                        num_found += 1
                if num_found >= num and num_found == len(tag_list):  # this means no other tags exist
                    f_out.write(work_id)
                    f_out.write("\n")
                    seen.append(work_id)
    f_in.close()
    f_out.close()


def get_status_ids(csv_in, out_txt_name, status):
    """
    Get all the ids of fics with a specific status
    :param csv_in: a csv created using ao3_get_fanfics.py
    :param out_txt_name: a txt file containing each id on a new line
    :param status: the status to check
    :return:
    """
    f_in = open(csv_in, 'r+')
    f_out = open(out_txt_name, 'w')
    reader = csv.reader(f_in)
    header = next(reader)
    # print(header)
    stat_names = ['work_id', 'title', 'rating', 'category', 'fandom', 'relationship', 'character', 'additional tags',
                  'language',
                  'published', 'status', 'status date', 'words', 'chapters', 'comments', 'kudos', 'bookmarks', 'hits',
                  'body']
    header = {stat: stat_names.index(stat) for stat in stat_names}
    seen = []
    for row in reader:
        if row[header['status']] == status:  # if the lists of categories are identical
            work_id = row[header['work_id']]
            if work_id in seen:
                print(out_txt_name, work_id)
            else:
                f_out.write(work_id)
                f_out.write("\n")
                seen.append(work_id)
    f_in.close()
    f_out.close()


def get_published_year_ids(csv_in, out_txt_name, year):
    """
    Get ids of all fics published in a given year
    :param csv_in: a csv created using ao3_get_fanfics.py
    :param out_txt_name: a txt file containing each id on a new line
    :param year: a year (can be a number or a string)
    :return:
    """
    if type(year) != str:
        year = str(year)
    f_in = open(csv_in, 'r+')
    f_out = open(out_txt_name, 'w')
    reader = csv.reader(f_in)
    header = next(reader)
    # print(header)
    stat_names = ['work_id', 'title', 'rating', 'category', 'fandom', 'relationship', 'character', 'additional tags',
                  'language',
                  'published', 'status', 'status date', 'words', 'chapters', 'comments', 'kudos', 'bookmarks', 'hits',
                  'body']
    header = {stat: stat_names.index(stat) for stat in stat_names}
    seen = []
    for row in reader:
        if year in row[header['published']]:  # if the lists of categories are identical
            work_id = row[header['work_id']]
            if work_id in seen:
                print(out_txt_name, work_id)
            else:
                f_out.write(work_id)
                f_out.write("\n")
                seen.append(work_id)
    f_in.close()
    f_out.close()


def store_stats(stat_dict, filename, stat_tup = ('title', 'rating', 'category', 'fandom', 'relationship', 'character', 'additional tags', 'published', 'status', 'status date', 'words', 'chapters', 'comments', 'kudos', 'bookmarks', 'hits')):
    """
    Store frequencies of various statistics (calculated using ao3_count_stats_from_csv) in the specified excel file
    :param stat_dict: a stat dict created using ao3_count_stats_from_csv
    :param filename: an excel file to store the result in
    :param stat_tup: the statistics to store information about
    :return:
    """
    wb = openpyxl.Workbook()
    ws = wb.active  # get a handle to the sheet in the workbook
    ws['A1'] = "AO3 stats"
    for col in range(1, len(stat_tup) * 2 + 1, 2):
        stat = stat_tup[(col - 1) // 2]
        print(stat, end="... ")
        entry = stat_dict[stat]
        entry_list = sorted(entry.items(), key=operator.itemgetter(1), reverse=True)
        ws.cell(row=2, column=col).value = stat
        for row in range(3, len(entry_list) + 3):  # this number is equal to the number of unique words with that suffix
            idx = row - 3
            value = entry_list[idx][0]
            try:
                ws.cell(row=row, column=col).value = value
            except openpyxl.utils.exceptions.IllegalCharacterError:
                value = input(value + " invalid. New value: ")
                ws.cell(row=row, column=col).value = value
            ws.cell(row=row, column=col+1).value = entry_list[idx][1]  # frequency
        if not filename.endswith(".xlsx"):
            wb.save(filename + ".xlsx")  # add the type extension if not included
        else:
            wb.save(filename)


def numpy_stats(csv_in, stat_list, out_csv_name, ratios=()):
    """
    Calculate mean, standard deviation, etc. of numerical stats from a fanfic csv and store the results in a csv file
    :param csv_in: a csv created using ao3_get_fanfics.py
    :param stat_list: a list of statistics to store information about
    :param out_csv_name: a csv to store the results in
    :param ratios: ratios to store information about. e.g., [("kudos", "hits"), ("comments", "hits")]
    :return:
    """
    # TODO: fix ratio stuff. Values aren't showing up right
    f_in = open(csv_in, 'r+')
    reader = csv.reader(f_in)
    header = next(reader)
    stat_names = ['work_id', 'title', 'rating', 'category', 'fandom', 'relationship', 'character', 'additional tags',
                  'language',
                  'published', 'status', 'status date', 'words', 'chapters', 'comments', 'kudos', 'bookmarks', 'hits',
                  'body']
    header = {stat: stat_names.index(stat) for stat in stat_names}
    stat_dict = {}
    seen = []
    for value in stat_list:
        stat_dict[value] = {"array": [], "samples": 0, "mean": 0, "median": 0, "min": 0, "max": 0, "std": 0}
    for ratio in ratios:
        stat_dict[ratio] = {"array": [], "samples": 0, "mean": 0, "median": 0, "min": 0, "max": 0, "std": 0}
    for row in reader:
        work_id = row[header["work_id"]]
        if work_id in seen:
            print(csv_in, work_id)
            continue
        for item in stat_list:
            value = row[header[item]]
            try:
                value = int(value)
            except ValueError:
                if value == "null" or value == "":
                    value = 0
                else:
                    print("ValueError:", value)
                    return
            stat_dict[item]["array"].append(value)
        for ratio in ratios:
            val0 = row[header[ratio[0]]]
            val1 = row[header[ratio[1]]]
            try:
                val0 = int(val0)
                val1 = int(val1)
            except ValueError:
                if val0 == "null" or val0 == "":
                    val0 = 0
                else:
                    val0 = int(val0)
                if val1 == "null" or val1 == "":
                    val1 = 0
                elif val0 != 0:  # if neither thing worked
                    print("ValueError:", val0, val1)
                    return
                else:
                    val1 = int(val1)
            try:
                ratio_val = val0 / val1
            except ZeroDivisionError:
                ratio_val = 0
            stat_dict[ratio]["array"].append(ratio_val)
            seen.append(work_id)
    f_in.close()
    # print(ratios)
    # print(stat_list)
    iterate_list = stat_list + ratios
    # for item in ratios:
    #     iterate_list.append(item)
    for item in iterate_list:  # iterate through both
        entry = stat_dict[item]
        entry["samples"] = len(entry["array"])  # easier to get len of list than get len of numpy array
        np_arr = np.array(entry["array"])
        entry["array"] = np_arr
        entry["mean"] = np.mean(np_arr)
        entry["median"] = np.median(np_arr)
        entry["min"] = np.amin(np_arr)
        entry["max"] = np.amax(np_arr)
        entry["std"] = np.std(np_arr)
    f_out = open(out_csv_name, "w")
    writer = csv.writer(f_out, delimiter='\t', quotechar='|')
    writer.writerow(["stat","samples","mean","median","minimum","maximum","standard deviation"])
    for value in stat_list:  # iterate through list rather than dict to preserve order
        entry = stat_dict[value]
        row = [value, entry["samples"], entry["mean"], entry["median"], entry["min"],
                                                    entry["max"], entry["std"]]
        writer.writerow(row)
    for ratio in ratios:
        entry = stat_dict[ratio]
        row = ["%s to %s" % (ratio[0], ratio[1]), entry["samples"], entry["mean"], entry["median"], entry["min"],
                                                    entry["max"], entry["std"]]
        writer.writerow(row)
    f_out.close()
    return stat_dict


def get_anomaly_info(stat_csv):
    """
    Analyze a csv created by numpy_stats to find statistical anomalies based on standard deviation
    :param stat_csv:
    :return:
    """
    f_in = open(stat_csv)
    reader = csv.reader(f_in, delimiter="\t", quotechar='|')
    header = next(reader)
    stat_conds = {}
    for row in reader:
        stat, samples, mean, median, min, max, std = row
        mean = float(mean)
        std = float(std)
        stat_conds[stat] = (mean-std, mean+std)  # find the abnormally low and high values based on mean and standard deviation
    f_in.close()
    return stat_conds


def get_anomaly_ids(stat_conds, csv_in, outfiledict, ratio_dict=()):
    """
    Get the ids of fics that are anomalous according to stat_conds
    :param stat_conds: maps each stat to the low threshold and the high threshold
    :param csv_in: a csv created using ao3_work_ids.py
    :param outfiledict: maps stats (e.g., "comments") to the filenames to store the info in. "val" will be replaced with "hi" or "lo"
    :return:
    """
    stat_names = ['work_id', 'title', 'rating', 'category', 'fandom', 'relationship', 'character',
                  'additional tags',
                  'language',
                  'published', 'status', 'status date', 'words', 'chapters', 'comments', 'kudos', 'bookmarks',
                  'hits',
                  'body']
    header = {stat: stat_names.index(stat) for stat in stat_names}
    seen = []
    for stat in stat_conds:
        if stat not in outfiledict:
            continue  # we'll get it in the ratios
        print(stat)
        min, max = stat_conds[stat]
        lo_name = outfiledict[stat].replace("val", "lo")
        hi_name = outfiledict[stat].replace("val", "hi")
        out_lo = open(lo_name, "w")
        out_hi = open(hi_name, "w")
        with open(csv_in) as f_in:
            reader = csv.reader(f_in)
            next(reader)
            for row in reader:
                work_id = row[header['work_id']]
                if work_id in seen:
                    print("anomaly", work_id)
                    continue
                value = row[header[stat]]
                try:
                    value = int(value)
                except ValueError:
                    value = 0
                if value < min:
                    out_lo.write(work_id + '\n')
                    seen.append(work_id)
                elif value > max:
                    out_hi.write(work_id + '\n')
                    seen.append(work_id)
        out_hi.close()
        out_lo.close()
    for ratio in ratio_dict:
        ratio_name = "%s to %s" % (ratio[0], ratio[1])
        print(ratio)
        min, max = stat_conds[ratio_name]
        lo_name = ratio_dict[ratio].replace("val", "lo")
        hi_name = ratio_dict[ratio].replace("val", "hi")
        out_lo = open(lo_name, "w")
        out_hi = open(hi_name, "w")
        with open(csv_in) as f_in:
            reader = csv.reader(f_in)
            next(reader)
            for row in reader:
                work_id = row[header['work_id']]
                if work_id in seen:
                    print("anomaly", work_id)
                    continue
                val0 = row[header[ratio[0]]]
                val1 = row[header[ratio[1]]]
                try:
                    val0 = int(val0)
                    val1 = int(val1)
                except ValueError:
                    if val0 == "null" or val0 == "":
                        val0 = 0
                    else:
                        val0 = int(val0)
                    if val1 == "null" or val1 == "":
                        val1 = 0
                    elif val0 != 0:  # if neither thing worked
                        print("ValueError:", val0, val1)
                        return
                    else:
                        val1 = int(val1)
                try:
                    ratio_val = val0 / val1
                except ZeroDivisionError:
                    ratio_val = 0
                if ratio_val < min:
                    out_lo.write(work_id + '\n')
                    seen.append(work_id)
                elif ratio_val > max:
                    out_hi.write(work_id + '\n')
                    seen.append(work_id)
        out_hi.close()
        out_lo.close()


def remove_ids_from_idlist(idlistfile, idfile_to_remove):
    to_remove = []
    f_in = open(idfile_to_remove)
    for line in f_in:
        line = line.strip()
        to_remove.append(line)
    f_in.close()
    f_in = open(idlistfile)
    outname = idlistfile.replace(".txt", "temp.txt")
    f_out = open(outname, "w")
    for line in f_in:
        if line.strip() not in to_remove:
            f_out.write(line)  # includes the newline
    f_out.close()
    f_in.close()
    os.rename(outname, idlistfile)


def only_english(csv_in, csv_out):
    """
    Create a new csv that only has fics in English
    :param csv_in: a csv created using ao3_get_fanfics.py
    :param csv_out: the csv to store the results in
    :return:
    """
    f_in = open(csv_in, 'r+')
    f_out = csv.writer(open(csv_out, 'a'), delimiter=',')
    reader = csv.reader(f_in)
    header = next(reader)
    langidx = header.index("language")
    for row in reader:
        print("only_english processing", row[0])
        lang = row[langidx]
        if lang == "English":
            f_out.writerow(row)
    f_in.close()


def category_ids(proj_dir, fandoms, categories):
    for fandom in fandoms:
        _fandom = fandom.replace(" ", "_")
        for category in categories:
            print(fandom, category)
            str_category = category.replace("/", "")
            get_category_ids(os.path.join(proj_dir, "CSV/%s_edit.csv" % _fandom), os.path.join(proj_dir, "Fanfic lists/%s %s.txt"
                    % (fandom, str_category)), [category])


def au_ids(proj_dir, fandoms):
    for fandom in fandoms:
        _fandom = fandom.replace(" ", "_")
        print(fandom, "AU")
        get_tag_ids(os.path.join(proj_dir, "CSV/%s_edit.csv" % _fandom), os.path.join(proj_dir, "Fanfic lists/%s AU.txt" % fandom),
                    ["Alternate Universe", "AU"], num=1, exact=False, solo=True)


def other_tags_ids(proj_dir, fandoms, tags):
    for fandom in fandoms:
        _fandom = fandom.replace(" ", "_")
        for tag in tags:
            print(fandom, tag)
            get_tag_ids(os.path.join(proj_dir, "CSV/%s_edit.csv" % _fandom),
                        os.path.join(proj_dir, "Fanfic lists/%s %s.txt" % (fandom, tag.replace("/", " "))), [tag], num=1, exact=False, solo=True)


def status_ids(proj_dir, fandoms, statuses):
    for fandom in fandoms:
        _fandom = fandom.replace(" ", "_")
        for status in statuses:
            print(fandom, status)
            get_status_ids(os.path.join(proj_dir, "CSV/%s_edit.csv" % _fandom),
                           os.path.join(proj_dir, "Fanfic lists/%s %s.txt" %(fandom, status)), status)


def year_ids(proj_dir, fandoms, years):
    for fandom in fandoms:
        _fandom = fandom.replace(" ", "_")
        for year in years:
            print(fandom, year)
            get_published_year_ids(os.path.join(proj_dir, "CSV/%s_edit.csv" % _fandom), os.path.join(proj_dir,
                    "Fanfic lists/%s %s.txt" % (fandom, year)), year)


def word_ids(proj_dir, fandoms, range_tuples):
    for fandom in fandoms:
        _fandom = fandom.replace(" ", "_")
        for rtuple in range_tuples:
            print(fandom, rtuple)
            get_numwords_ids(os.path.join(proj_dir, "CSV/%s_edit.csv" %_fandom),
                         os.path.join(proj_dir, "Fanfic lists/%s %d-%d.txt" % (fandom, rtuple[0], rtuple[1])), rtuple[0], rtuple[1])


def rating_ids(proj_dir, fandoms, ratings):
    for fandom in fandoms:
        _fandom = fandom.replace(" ", "_")
        for rating in ratings:
            print(fandom, rating)
            get_rating_ids(os.path.join(proj_dir, "CSV/%s_edit.csv" % _fandom),
                   os.path.join(proj_dir, "Fanfic lists/%s %s.txt" % (fandom, rating)), rating)


def fandom_id_files(proj_dir):
    print("DW")
    get_fandom_ids(os.path.join(proj_dir, "CSV/Doctor_Who_edit.csv"), os.path.join(proj_dir, "Fanfic lists/New Who.txt"),
                   ["Doctor Who", "Doctor Who (2005)", "Doctor Who & Related Fandoms"])
    get_fandom_ids(os.path.join(proj_dir, "CSV/Doctor_Who_edit.csv"), os.path.join(proj_dir, "Fanfic lists/Classic Who.txt"),
                   ["Doctor Who", "Doctor Who (1963)", "Doctor Who: Eighth Doctor Adventures - Various Authors", "Doctor Who & Related Fandoms"])
    print("H")
    get_fandom_ids(os.path.join(proj_dir, "CSV/Hamilton_edit.csv"), os.path.join(proj_dir, "Fanfic lists/Hamilton.txt"),
                   ["Hamilton - Miranda", "Hamilton - Fandom", "Historical RPF", "18th & 19th Century CE RPF", "American Revolution RPF",
                    "Hamilton-Miranda", "18th Century CE RPF", "19th Century CE RPF", "Hamilton- Miranda", "Alexander Hamilton - Ron Chernow"])
    print("LM")
    get_fandom_ids(os.path.join(proj_dir, "CSV/Les_Mis_edit.csv"), os.path.join(proj_dir, "Fanfic lists/Les Mis.txt"),
                   ["Les Miserables - All Media Types", "Les Miserables - Victor Hugo", "Les Miserables (2012)",
                    "Les Miserables - Schonberg/Boublil", "Les Miserables", "Les Miserables (Dallas 2014)", "les mis"])
    print("SH")
    get_fandom_ids(os.path.join(proj_dir, "CSV/Sherlock_edit.csv"), os.path.join(proj_dir, "Fanfic lists/Sherlock.txt"),
                   ["Sherlock (TV)", "Sherlock Holmes & Related Fandoms", "Elementary (TV)", "Elementary", "Sherlock Holmes - Arthur Conan Doyle",
                    "Sherlock - Fandom", "Sherlock BBC", "BBC Sherlock", "Sherlock Holmes (2009)", "Sherlock Holmes (Downey films)",
                    "Sherlock Holmes (1984 TV)", "Sherlock Holmes - Doyle"])
    print("BBC SH")
    get_fandom_ids(os.path.join(proj_dir, "CSV/Sherlock_edit.csv"), os.path.join(proj_dir, "Fanfic lists/BBC Sherlock.txt"),
                   ["Sherlock (TV)", "Sherlock Holmes & Related Fandoms", "Sherlock - Fandom", "Sherlock BBC", "BBC Sherlock"])
    print("ST")
    get_fandom_ids(os.path.join(proj_dir, "CSV/Star_Trek_edit.csv"), os.path.join(proj_dir, "Fanfic lists/Star Trek.txt"), ["Star Trek: Alternate Original Series (Movies)", "Star Trek (2009)", "Star Trek: Voyager", "Star Trek: Enterprise", "Star Trek: The Original Series", "Star Trek: Deep Space Nine", "Star Trek: The Next Generation", "Star Trek",
    "Star Trek: Discovery", "Star Trek XI", "Star Trek: 2009", "Star Trek: Mirror Universe", "Star Trek Voyager",
        "StarTrek: Voyager", "Star Trek Into Darkness - Fandom", "Star Trek: Into Darkness - Fandom", "Star Trek Reboot",
        "Star Trek AOS", "Star Trek Enterprise", "StarTrek: Enterprise", "Star Trek - Various Authors", "Star Trek Online",
        "ST:AOS - Fandom", "StarTrek: The Next Generation", "Star Trek Into Darkness - Fandom", "Star Trek The Next Generation"])
    print("T")
    get_fandom_ids(os.path.join(proj_dir, "CSV/Tolkien_edit.csv"), os.path.join(proj_dir, "Fanfic lists/Tolkien.txt"), ["The Hobbit - All Media Types", "The Silmarillion and other histories of Middle-Earth - J. R. R. Tolkien",
        "The Lord of the Rings - J. R. R. Tolkien", "The Hobbit (Jackson Movies)", "The Hobbit - J. R. R. Tolkien",
        "The Lord of the Rings - All Media Types", "The Hobbit (2012)", "TOLKIEN J. R. R. - Works & Related Fandoms",
        "Lord of the Rings (2001 2002 2003)", "The Lord of the Rings (Movies)", "Lord of the Rings - Fandom", "Lord of the Rings - Tolkien",
        "The Hobbit" ])
    print("U")
    get_fandom_ids(os.path.join(proj_dir, "CSV/Undertale_edit.csv"), os.path.join(proj_dir, "Fanfic lists/Undertale.txt"), ["Undertale (Video Game)", "Undertale", "Underfell - Fandom", "Undertail - Fandom", "Undertail - Fandom", "underswap"])


def csv_id_files(proj_dir, fandoms):
    for fandom in fandoms:
        _fandom = fandom.replace(" ", "_")
        get_csv_ids(os.path.join(proj_dir, "CSV/%s_edit.csv" % _fandom),
                   os.path.join(proj_dir, "Fanfic lists/%s all.txt" % fandom))


def get_fandom_group_id_files(proj_dir):
    # TODO: figure out why this isn't working
    print("Wholock")
    get_fandom_group_ids(os.path.join(proj_dir, "CSV/Doctor_Who_edit.csv"), os.path.join(proj_dir, "Fanfic lists/Wholock.txt"),
                         [["Doctor Who", "Doctor Who (2005)", "Doctor Who & Related Fandoms", "Doctor Who (1963)", "Doctor Who: Eighth Doctor Adventures - Various Authors"],
                          ["Sherlock (TV)", "Sherlock Holmes & Related Fandoms", "Sherlock - Fandom", "Sherlock BBC", "BBC Sherlock"]])
    print("SuperWhoLock")
    get_fandom_group_ids(os.path.join(proj_dir, "CSV/Doctor_Who_edit.csv"), os.path.join(proj_dir, "Fanfic lists/SuperWhoLock.txt"),
                         [["Doctor Who", "Doctor Who (2005)", "Doctor Who & Related Fandoms", "Doctor Who (1963)", "Doctor Who: Eighth Doctor Adventures - Various Authors"],
                          ["Sherlock (TV)", "Sherlock Holmes & Related Fandoms", "Sherlock - Fandom", "Sherlock BBC", "BBC Sherlock"],
                          ["Supernatural"]])


def fandom_numpy_stats(proj_dir, fandoms):
    stat_list = ["words", "comments", "kudos", "hits", "bookmarks"]
    ratios = [("comments", "hits"), ("kudos", "hits"), ("bookmarks", "hits"), ("comments", "kudos")]
    for fandom in fandoms:
        print(fandom)
        numpy_stats(os.path.join(proj_dir, "CSV/%s_edit.csv" % fandom.replace(" ", "_")), stat_list, os.path.join(proj_dir,
            "CSV stats/%s num stats.csv" % fandom), ratios)


def fandom_anomaly_hist(proj_dir):
    stat_tup = ("words", "comments", "kudos", "hits", "bookmarks")
    fandoms = ("Doctor Who", "Les Mis", "Hamilton", "Sherlock", "Star Trek", "Tolkien", "Undertale")
    for fandom in fandoms:
        print(fandom)
        np_stats = numpy_stats(os.path.join(proj_dir, "CSV/%s_edit.csv" % fandom.replace(" ", "_")), stat_tup, os.path.join(proj_dir,
                                                                    "CSV stats/%s num stats.csv" % fandom))
        anomaly_hist(np_stats)


def fandom_anomaly_ids(proj_dir, fandoms):
    stat_list = ["words", "comments", "kudos", "hits", "bookmarks"]
    ratios = [("comments", "hits"), ("kudos", "hits"), ("bookmarks", "hits"), ("comments", "kudos")]
    for fandom in fandoms:
        _fandom = fandom.replace(" ", "_")
        print(fandom)
        stat_conds = get_anomaly_info(os.path.join(proj_dir, "CSV stats/%s num stats.csv" % fandom))
        outfiledict = {stat: os.path.join(proj_dir, "Fanfic lists1/%s %s val.txt" % (fandom, stat)) for stat in stat_list}  # TODO: change back
        ratio_dict = {ratio: os.path.join(proj_dir, "Fanfic lists1/%s %s to %s val.txt" % (fandom, ratio[0], ratio[1])) for ratio in ratios}
        # TODO: change back
        get_anomaly_ids(stat_conds, os.path.join(proj_dir, "CSV/%s_edit.csv" % _fandom), outfiledict, ratio_dict=ratio_dict)


def duplicate_ids(proj_dir, idfiles):
    """
    Find the duplicate ids across multiple id files
    :param proj_dir: directory of the project
    :param idfiles: a list of id files
    :return: a list of the duplicate ids
    """
    duplicates = []
    seen = []
    for filename in idfiles:
        f_in = open(os.path.join(proj_dir, "Fanfic lists/" + filename))
        for line in f_in:
            current_id = line.strip()
            if current_id in seen:
                if current_id not in duplicates:
                    duplicates.append(current_id)
            else:
                seen.append(current_id)
        f_in.close()
    return duplicates


def unique_ids(proj_dir, idfiles):
    """
    Find the ids in the first file that are not in the other files
    :param proj_dir: directory of the project
    :param idfiles: a list of id files
    :return: a list of the unique ids
    """
    first_list = []
    f_in = open(os.path.join(proj_dir, "Fanfic lists/" + idfiles[0]))
    for line in f_in:
        current_id = line.strip()
        first_list.append(current_id)
    f_in.close()
    for filename in idfiles[1:]:
        f_in = open(os.path.join(proj_dir, "Fanfic lists/" + filename))
        for line in f_in:
            current_id = line.strip()
            if current_id in first_list:
                first_list.remove(current_id)
        f_in.close()
    return first_list


proj_dir = "/Volumes/2TB/Final_project"
fandoms = ("Doctor Who", "Hamilton", "Les Mis", "Sherlock", "Star Trek", "Tolkien", "Undertale")
categories = ("F/M", "M/M", "F/F", "Gen", "Multi", "Other")
tags = ("Fluff", "Angst", "Humor", "Romance", "Hurt/Comfort", "Established", "Friendship", "Crack")
statuses = ("Completed", "Updated")
years = [str(x) for x in range(2009, 2019)]
range_tuples = [(1, 100), (1, 1000), (1001, 5000), (5001, 10000), (1001, 10000), (10001, 50000), (50001, 100000), (10001, 100000),
                   (100001, 500000), (500001, 1000000), (100001, 1000000)]
ratings = ("General Audiences", "Teen And Up Audiences", "Mature", "Explicit", "Not Rated")

fix_fands = ["Doctor Who", "Hamilton", "Tolkien", "Undertale"]
# for fandom in fix_fands:
#     print(fandom)
#     _fandom = fandom.replace(" ", "_")
#     ao3_edit_csv(os.path.join(proj_dir, "CSV/%s_works.csv" % _fandom), os.path.join(proj_dir, "CSV/%s_edit.csv"))

# csv_id_files(proj_dir, fix_fands)
# category_ids(proj_dir, fix_fands, categories)
# au_ids(proj_dir, fix_fands)
# other_tags_ids(proj_dir, fix_fands, tags)
# status_ids(proj_dir, fix_fands, statuses)
# year_ids(proj_dir, fix_fands, years)
# word_ids(proj_dir, fix_fands, range_tuples)
# rating_ids(proj_dir, fix_fands, ratings)
# fandom_id_files(proj_dir)
fandom_numpy_stats(proj_dir, fandoms)
fandom_anomaly_ids(proj_dir, fandoms)

# for filename in os.listdir(os.path.join(proj_dir, "Fanfic lists")):
#     if filename.endswith(".txt"):
#         print(filename, end=" ")
#         dup_name = os.path.join(proj_dir, "Duplicates/%s" % filename)
#         if os.path.isfile(dup_name):
#             print("duplicates")
#             remove_ids_from_idlist(os.path.join(proj_dir, "Fanfic lists/%s" % filename), dup_name)
#         else:
#             print()

# fandom_numpy_stats(proj_dir, fandoms)
# fandom_anomaly_ids(proj_dir, fandoms)

# for fandom in fandoms:
#     print(fandom)
#     uniques = unique_ids(proj_dir, ["%s 10000-100000.txt" % fandom, "%s 1001-10000.txt" % fandom])
#     f_out = open(os.path.join(proj_dir, "Fanfic lists/%s 10001-100000.txt" % fandom), "w")
#     for item in uniques:
#         f_out.write(item + "\n")
#     f_out.close()

# duplicate_ids(proj_dir, ["Doctor Who 5001-10000.txt", "Doctor Who 10000-100000.txt"])
# unique_ids(proj_dir, ["Doctor Who_hits_hi.txt", "Doctor Who_kudos_hi.txt"])

"""
/Volumes/2TB/Final_project/Fanfic lists/Doctor Who FM.txt 12020949
/Volumes/2TB/Final_project/Fanfic lists/Doctor Who FM.txt 5895805
/Volumes/2TB/Final_project/Fanfic lists/Doctor Who FM.txt 1194534
/Volumes/2TB/Final_project/Fanfic lists/Doctor Who FM.txt 1131479
/Volumes/2TB/Final_project/Fanfic lists/Doctor Who FM.txt 1105838
/Volumes/2TB/Final_project/Fanfic lists/Doctor Who FM.txt 3765820
/Volumes/2TB/Final_project/Fanfic lists/Doctor Who Gen.txt 7663738
/Volumes/2TB/Final_project/Fanfic lists/Doctor Who Gen.txt 1189287
/Volumes/2TB/Final_project/Fanfic lists/Doctor Who Gen.txt 1190265
/Volumes/2TB/Final_project/Fanfic lists/Doctor Who Gen.txt 966471
/Volumes/2TB/Final_project/Fanfic lists/Hamilton FM.txt 13352475
/Volumes/2TB/Final_project/Fanfic lists/Hamilton FM.txt 13349595
/Volumes/2TB/Final_project/Fanfic lists/Hamilton FM.txt 13347183
/Volumes/2TB/Final_project/Fanfic lists/Hamilton FM.txt 13241754
/Volumes/2TB/Final_project/Fanfic lists/Hamilton FM.txt 13327818
/Volumes/2TB/Final_project/Fanfic lists/Hamilton FM.txt 13165368
/Volumes/2TB/Final_project/Fanfic lists/Hamilton FM.txt 12306912
/Volumes/2TB/Final_project/Fanfic lists/Hamilton FM.txt 12292116
/Volumes/2TB/Final_project/Fanfic lists/Hamilton FM.txt 10424598
/Volumes/2TB/Final_project/Fanfic lists/Hamilton FM.txt 12246579
/Volumes/2TB/Final_project/Fanfic lists/Hamilton FM.txt 13304880
/Volumes/2TB/Final_project/Fanfic lists/Hamilton MM.txt 11883954
/Volumes/2TB/Final_project/Fanfic lists/Hamilton MM.txt 10363875
/Volumes/2TB/Final_project/Fanfic lists/Hamilton MM.txt 13353165
/Volumes/2TB/Final_project/Fanfic lists/Hamilton MM.txt 13276701
/Volumes/2TB/Final_project/Fanfic lists/Hamilton MM.txt 10353570
/Volumes/2TB/Final_project/Fanfic lists/Hamilton MM.txt 13342140
/Volumes/2TB/Final_project/Fanfic lists/Hamilton MM.txt 13252596
/Volumes/2TB/Final_project/Fanfic lists/Hamilton MM.txt 13349766
/Volumes/2TB/Final_project/Fanfic lists/Hamilton MM.txt 13348368
/Volumes/2TB/Final_project/Fanfic lists/Hamilton MM.txt 13178529
/Volumes/2TB/Final_project/Fanfic lists/Hamilton MM.txt 13342461
/Volumes/2TB/Final_project/Fanfic lists/Hamilton MM.txt 12575128
/Volumes/2TB/Final_project/Fanfic lists/Hamilton MM.txt 13282836
/Volumes/2TB/Final_project/Fanfic lists/Hamilton MM.txt 13323093
/Volumes/2TB/Final_project/Fanfic lists/Hamilton MM.txt 13284711
/Volumes/2TB/Final_project/Fanfic lists/Hamilton MM.txt 13340388
/Volumes/2TB/Final_project/Fanfic lists/Hamilton MM.txt 12379977
/Volumes/2TB/Final_project/Fanfic lists/Hamilton MM.txt 13338951
/Volumes/2TB/Final_project/Fanfic lists/Hamilton MM.txt 13319781
/Volumes/2TB/Final_project/Fanfic lists/Hamilton MM.txt 13336203
/Volumes/2TB/Final_project/Fanfic lists/Hamilton MM.txt 12628113
/Volumes/2TB/Final_project/Fanfic lists/Hamilton MM.txt 13334211
/Volumes/2TB/Final_project/Fanfic lists/Hamilton MM.txt 13328703
/Volumes/2TB/Final_project/Fanfic lists/Hamilton MM.txt 11527713
/Volumes/2TB/Final_project/Fanfic lists/Hamilton MM.txt 13319922
/Volumes/2TB/Final_project/Fanfic lists/Hamilton MM.txt 13128351
/Volumes/2TB/Final_project/Fanfic lists/Hamilton MM.txt 13301202
/Volumes/2TB/Final_project/Fanfic lists/Hamilton MM.txt 12354762
/Volumes/2TB/Final_project/Fanfic lists/Hamilton MM.txt 13323795
/Volumes/2TB/Final_project/Fanfic lists/Hamilton MM.txt 10436811
/Volumes/2TB/Final_project/Fanfic lists/Hamilton MM.txt 12994017
/Volumes/2TB/Final_project/Fanfic lists/Hamilton MM.txt 12388026
/Volumes/2TB/Final_project/Fanfic lists/Hamilton MM.txt 13319577
/Volumes/2TB/Final_project/Fanfic lists/Hamilton MM.txt 13317909
/Volumes/2TB/Final_project/Fanfic lists/Hamilton MM.txt 12865263
/Volumes/2TB/Final_project/Fanfic lists/Hamilton MM.txt 13304235
/Volumes/2TB/Final_project/Fanfic lists/Hamilton MM.txt 9378599
/Volumes/2TB/Final_project/Fanfic lists/Hamilton MM.txt 6761590
/Volumes/2TB/Final_project/Fanfic lists/Hamilton FF.txt 13239381
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Gen.txt 12917652
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Gen.txt 9708923
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Gen.txt 13326354
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Gen.txt 10818699
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Gen.txt 13311780
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Gen.txt 8573068
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Gen.txt 13310295
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Multi.txt 13350303
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Multi.txt 13332426
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Multi.txt 13293231
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Multi.txt 13308705
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Multi.txt 12921027
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Multi.txt 11265828
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Multi.txt 13324317
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Multi.txt 9851060
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Other.txt 13330458
/Volumes/2TB/Final_project/Fanfic lists/Tolkien FM.txt 12422925
/Volumes/2TB/Final_project/Fanfic lists/Tolkien FM.txt 3778871
/Volumes/2TB/Final_project/Fanfic lists/Tolkien MM.txt 6294298
/Volumes/2TB/Final_project/Fanfic lists/Tolkien MM.txt 4555053
/Volumes/2TB/Final_project/Fanfic lists/Tolkien MM.txt 1129347
/Volumes/2TB/Final_project/Fanfic lists/Tolkien MM.txt 697748
/Volumes/2TB/Final_project/Fanfic lists/Tolkien MM.txt 1165274
/Volumes/2TB/Final_project/Fanfic lists/Undertale FM.txt 12788001
/Volumes/2TB/Final_project/Fanfic lists/Undertale Gen.txt 11476317
/Volumes/2TB/Final_project/Fanfic lists/Undertale Gen.txt 8920603
/Volumes/2TB/Final_project/Fanfic lists/Undertale Gen.txt 5129651
/Volumes/2TB/Final_project/Fanfic lists/Undertale Gen.txt 8501281
/Volumes/2TB/Final_project/Fanfic lists/Undertale Gen.txt 5607715
/Volumes/2TB/Final_project/Fanfic lists/Undertale Multi.txt 8794159
/Volumes/2TB/Final_project/Fanfic lists/Undertale Other.txt 9136294
/Volumes/2TB/Final_project/Fanfic lists/Doctor Who Completed.txt 12020949
/Volumes/2TB/Final_project/Fanfic lists/Doctor Who Completed.txt 7663738
/Volumes/2TB/Final_project/Fanfic lists/Doctor Who Completed.txt 5895805
/Volumes/2TB/Final_project/Fanfic lists/Doctor Who Completed.txt 2008722
/Volumes/2TB/Final_project/Fanfic lists/Doctor Who Completed.txt 2085354
/Volumes/2TB/Final_project/Fanfic lists/Doctor Who Completed.txt 1773139
/Volumes/2TB/Final_project/Fanfic lists/Doctor Who Completed.txt 1189287
/Volumes/2TB/Final_project/Fanfic lists/Doctor Who Completed.txt 1190265
/Volumes/2TB/Final_project/Fanfic lists/Doctor Who Completed.txt 1194534
/Volumes/2TB/Final_project/Fanfic lists/Doctor Who Completed.txt 1075635
/Volumes/2TB/Final_project/Fanfic lists/Doctor Who Completed.txt 966471
/Volumes/2TB/Final_project/Fanfic lists/Doctor Who Completed.txt 966368
/Volumes/2TB/Final_project/Fanfic lists/Doctor Who Completed.txt 930375
/Volumes/2TB/Final_project/Fanfic lists/Doctor Who Completed.txt 3765820
/Volumes/2TB/Final_project/Fanfic lists/Doctor Who Completed.txt 262635
/Volumes/2TB/Final_project/Fanfic lists/Doctor Who Updated.txt 1131479
/Volumes/2TB/Final_project/Fanfic lists/Doctor Who Updated.txt 1105838
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Completed.txt 10363875
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Completed.txt 10086296
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Completed.txt 13353261
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Completed.txt 13352475
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Completed.txt 13342140
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Completed.txt 13350303
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Completed.txt 13332426
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Completed.txt 13349766
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Completed.txt 13348368
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Completed.txt 13347183
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Completed.txt 13341234
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Completed.txt 13282836
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Completed.txt 13327818
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Completed.txt 13340388
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Completed.txt 13338951
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Completed.txt 13336203
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Completed.txt 13335636
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Completed.txt 13334211
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Completed.txt 13332942
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Completed.txt 10570644
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Completed.txt 13328703
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Completed.txt 13330458
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Completed.txt 7963879
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Completed.txt 13326354
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Completed.txt 11981517
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Completed.txt 10818699
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Completed.txt 12354762
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Completed.txt 13324317
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Completed.txt 13323795
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Completed.txt 13322394
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Completed.txt 13319577
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Completed.txt 13319385
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Completed.txt 13318668
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Completed.txt 13317909
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Completed.txt 9253997
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Completed.txt 13311780
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Completed.txt 13310295
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Completed.txt 13304235
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Completed.txt 13304880
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Completed.txt 13308402
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Completed.txt 7925812
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Completed.txt 7901473
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Completed.txt 6761590
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 11883954
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 12281412
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 12926160
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 13353165
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 13352811
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 13255794
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 13276701
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 10353570
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 13314087
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 12917652
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 13181403
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 13350456
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 7947316
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 13350000
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 13252596
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 13349595
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 13348206
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 13281603
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 12474668
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 13178529
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 13344735
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 13343205
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 13287819
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 13293231
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 13343364
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 13308705
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 13304415
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 13217142
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 13342461
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 13241754
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 13203054
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 12575128
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 10661775
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 13165368
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 13323093
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 13284711
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 12379977
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 13284573
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 13319781
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 10022444
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 12628113
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 13239381
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 10963401
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 9708923
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 12921027
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 6642676
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 11265828
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 12306912
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 5926771
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 13223403
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 11527713
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 13286346
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 13319922
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 13128351
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 12292116
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 13301202
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 13326495
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 13324881
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 11748897
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 11934168
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 10436811
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 13212894
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 12994017
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 11888109
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 13092189
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 12474268
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 12388026
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 13321014
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 13184589
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 9851060
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 13313868
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 10424598
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 8573068
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 12246579
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 12865263
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 13305450
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 13270842
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 13249860
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 11194281
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 12423564
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 12971052
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 9378599
/Volumes/2TB/Final_project/Fanfic lists/Hamilton Updated.txt 7681768
/Volumes/2TB/Final_project/Fanfic lists/Tolkien Completed.txt 10285889
/Volumes/2TB/Final_project/Fanfic lists/Tolkien Completed.txt 4555053
/Volumes/2TB/Final_project/Fanfic lists/Tolkien Completed.txt 3778871
/Volumes/2TB/Final_project/Fanfic lists/Tolkien Completed.txt 1129347
/Volumes/2TB/Final_project/Fanfic lists/Tolkien Completed.txt 697748
/Volumes/2TB/Final_project/Fanfic lists/Tolkien Completed.txt 1165274
/Volumes/2TB/Final_project/Fanfic lists/Tolkien Completed.txt 3762620
/Volumes/2TB/Final_project/Fanfic lists/Tolkien Updated.txt 12422925
/Volumes/2TB/Final_project/Fanfic lists/Tolkien Updated.txt 6294298
/Volumes/2TB/Final_project/Fanfic lists/Tolkien Updated.txt 3065678
/Volumes/2TB/Final_project/Fanfic lists/Undertale Completed.txt 12788001
/Volumes/2TB/Final_project/Fanfic lists/Undertale Completed.txt 11476317
/Volumes/2TB/Final_project/Fanfic lists/Undertale Completed.txt 8794159
/Volumes/2TB/Final_project/Fanfic lists/Undertale Completed.txt 9136294
/Volumes/2TB/Final_project/Fanfic lists/Undertale Completed.txt 8920603
/Volumes/2TB/Final_project/Fanfic lists/Undertale Completed.txt 8813101
/Volumes/2TB/Final_project/Fanfic lists/Undertale Completed.txt 8501281
/Volumes/2TB/Final_project/Fanfic lists/Undertale Completed.txt 5607715
/Volumes/2TB/Final_project/Fanfic lists/Undertale Updated.txt 5129651
/Volumes/2TB/Final_project/Fanfic lists/Undertale Updated.txt 6319888
/Volumes/2TB/Final_project/Fanfic lists/Undertale Updated.txt 5665993
/Volumes/2TB/Final_project/Fanfic lists/Doctor Who 2011.txt 262635
/Volumes/2TB/Final_project/Fanfic lists/Doctor Who 2012.txt 3765820
/Volumes/2TB/Final_project/Fanfic lists/Doctor Who 2013.txt 1105838
/Volumes/2TB/Final_project/Fanfic lists/Doctor Who 2013.txt 1075635
/Volumes/2TB/Final_project/Fanfic lists/Doctor Who 2013.txt 966471
/Volumes/2TB/Final_project/Fanfic lists/Doctor Who 2013.txt 966368
/Volumes/2TB/Final_project/Fanfic lists/Doctor Who 2013.txt 930375
/Volumes/2TB/Final_project/Fanfic lists/Doctor Who 2014.txt 2008722
/Volumes/2TB/Final_project/Fanfic lists/Doctor Who 2014.txt 2085354
/Volumes/2TB/Final_project/Fanfic lists/Doctor Who 2014.txt 1773139
/Volumes/2TB/Final_project/Fanfic lists/Doctor Who 2014.txt 1189287
/Volumes/2TB/Final_project/Fanfic lists/Doctor Who 2014.txt 1190265
/Volumes/2TB/Final_project/Fanfic lists/Doctor Who 2014.txt 1194534
/Volumes/2TB/Final_project/Fanfic lists/Doctor Who 2014.txt 1131479
/Volumes/2TB/Final_project/Fanfic lists/Doctor Who 2016.txt 7663738
/Volumes/2TB/Final_project/Fanfic lists/Doctor Who 2016.txt 5895805
/Volumes/2TB/Final_project/Fanfic lists/Doctor Who 2017.txt 12020949
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2016.txt 7947316
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2016.txt 6642676
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2016.txt 5926771
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2016.txt 7963879
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2016.txt 8573068
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2016.txt 7925812
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2016.txt 7901473
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2016.txt 7681768
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2016.txt 6761590
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2017.txt 11883954
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2017.txt 10363875
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2017.txt 10086296
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2017.txt 12281412
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2017.txt 12926160
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2017.txt 10353570
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2017.txt 12917652
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2017.txt 13181403
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2017.txt 12474668
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2017.txt 13217142
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2017.txt 13203054
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2017.txt 12575128
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2017.txt 10661775
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2017.txt 13165368
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2017.txt 12379977
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2017.txt 10022444
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2017.txt 12628113
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2017.txt 10963401
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2017.txt 9708923
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2017.txt 10570644
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2017.txt 12921027
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2017.txt 11265828
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2017.txt 12306912
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2017.txt 11527713
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2017.txt 13128351
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2017.txt 12292116
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2017.txt 11981517
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2017.txt 10818699
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2017.txt 12354762
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2017.txt 11748897
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2017.txt 11934168
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2017.txt 10436811
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2017.txt 13212894
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2017.txt 12994017
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2017.txt 11888109
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2017.txt 13092189
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2017.txt 12474268
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2017.txt 12388026
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2017.txt 13184589
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2017.txt 9851060
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2017.txt 9253997
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2017.txt 10424598
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2017.txt 12246579
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2017.txt 12865263
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2017.txt 11194281
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2017.txt 12423564
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2017.txt 12971052
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2017.txt 9378599
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13353165
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13353261
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13352811
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13255794
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13276701
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13352475
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13342140
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13314087
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13350456
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13350303
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13350000
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13332426
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13252596
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13349766
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13349595
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13348368
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13348206
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13281603
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13347183
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13178529
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13344735
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13343205
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13287819
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13293231
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13343364
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13308705
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13304415
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13342461
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13241754
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13341234
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13282836
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13327818
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13323093
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13284711
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13340388
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13338951
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13284573
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13319781
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13336203
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13335636
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13334211
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13239381
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13332942
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13328703
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13223403
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13286346
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13330458
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13319922
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13301202
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13326495
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13326354
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13324881
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13324317
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13323795
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13322394
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13321014
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13319577
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13319385
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13318668
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13317909
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13313868
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13311780
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13310295
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13305450
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13270842
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13304235
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13304880
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13249860
/Volumes/2TB/Final_project/Fanfic lists/Hamilton 2018.txt 13308402
/Volumes/2TB/Final_project/Fanfic lists/Tolkien 2013.txt 697748
/Volumes/2TB/Final_project/Fanfic lists/Tolkien 2014.txt 3065678
/Volumes/2TB/Final_project/Fanfic lists/Tolkien 2014.txt 1129347
/Volumes/2TB/Final_project/Fanfic lists/Tolkien 2015.txt 4555053
/Volumes/2TB/Final_project/Fanfic lists/Tolkien 2015.txt 3778871
/Volumes/2TB/Final_project/Fanfic lists/Tolkien 2015.txt 3762620
/Volumes/2TB/Final_project/Fanfic lists/Tolkien 2016.txt 6294298
/Volumes/2TB/Final_project/Fanfic lists/Tolkien 2017.txt 12422925
/Volumes/2TB/Final_project/Fanfic lists/Tolkien 2017.txt 10285889
/Volumes/2TB/Final_project/Fanfic lists/Undertale 2015.txt 5129651
/Volumes/2TB/Final_project/Fanfic lists/Undertale 2016.txt 8794159
/Volumes/2TB/Final_project/Fanfic lists/Undertale 2016.txt 8920603
/Volumes/2TB/Final_project/Fanfic lists/Undertale 2016.txt 8813101
/Volumes/2TB/Final_project/Fanfic lists/Undertale 2016.txt 8501281
/Volumes/2TB/Final_project/Fanfic lists/Undertale 2016.txt 6319888
/Volumes/2TB/Final_project/Fanfic lists/Undertale 2016.txt 5665993
/Volumes/2TB/Final_project/Fanfic lists/Undertale 2016.txt 5607715
/Volumes/2TB/Final_project/Fanfic lists/Undertale 2017.txt 12788001
/Volumes/2TB/Final_project/Fanfic lists/Undertale 2017.txt 11476317
/Volumes/2TB/Final_project/Fanfic lists/Undertale 2017.txt 9136294
"""

# New Who: ["Doctor Who", "Doctor Who (2005)", "Doctor Who & Related Fandoms"]
    # Hamilton: ["Hamilton - Miranda", "Hamilton - Fandom", "Historical RPF", "18th & 19th Century CE RPF", "American Revolution RPF"
    # "Hamilton-Miranda", "18th Century CE RPF", "19th Century CE RPF", "Hamilton- Miranda", "Alexander Hamilton - Ron Chernow"]
    # Les Mis: ["Les Miserables - All Media Types", "Les Miserables - Victor Hugo", "Les Miserables (2012)", "Les Miserables - Schonberg/Boublil",
    # "Les Miserables", "Les Miserables (Dallas 2014)", "les mis"]
    # Sherlock: ["Sherlock (TV)", "Sherlock Holmes & Related Fandoms", "Elementary (TV)", "Elementary", "Sherlock Holmes - Arthur Conan Doyle",
    # "Sherlock - Fandom", "Sherlock BBC", "BBC Sherlock", "Sherlock Holmes (2009)", "Sherlock Holmes (Downey films)",
    # "Sherlock Holmes (1984 TV)", "Sherlock Holmes - Doyle"]
    # Star Trek: ["Star Trek: Alternate Original Series (Movies)", "Star Trek (2009)", "Star Trek: Voyager", "Star Trek: Enterprise",
    # "Star Trek: The Original Series", "Star Trek: Deep Space Nine", "Star Trek: The Next Generation", "Star Trek",
    # "Star Trek: Discovery", "Star Trek XI", "Star Trek: 2009", "Star Trek: Mirror Universe", "Star Trek Voyager",
    # "StarTrek: Voyager", "Star Trek Into Darkness - Fandom", "Star Trek: Into Darkness - Fandom", "Star Trek Reboot",
    # "Star Trek AOS", "Star Trek Enterprise", "StarTrek: Enterprise", "Star Trek - Various Authors", "Star Trek Online",
    # "ST:AOS - Fandom", "StarTrek: The Next Generation", "Star Trek Into Darkness - Fandom", "Star Trek The Next Generation"]
    # Tolkien: ["The Hobbit - All Media Types", "The Silmarillion and other histories of Middle-Earth - J. R. R. Tolkien",
    # "The Lord of the Rings - J. R. R. Tolkien", "The Hobbit (Jackson Movies)", "The Hobbit - J. R. R. Tolkien",
    # "The Lord of the Rings - All Media Types", "The Hobbit (2012)", "TOLKIEN J. R. R. - Works & Related Fandoms",
    # "Lord of the Rings (2001 2002 2003)", "The Lord of the Rings (Movies)", "Lord of the Rings - Fandom", "Lord of the Rings - Tolkien",
    # "The Hobbit" ]
    # Undertale: ["Undertale (Video Game)", "Undertale", "Underfell - Fandom", "Undertail - Fandom", "Undertail - Fandom", "underswap"]