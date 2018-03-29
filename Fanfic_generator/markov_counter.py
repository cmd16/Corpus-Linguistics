import openpyxl
verbose = True

#Note: deal with contractions

def remove_chars(array):
    if verbose:
        print("removing chars")
    for char in "'\".,;-_[]{}\\/+=|`~<>?!()*&^%$#@\"0123456789£":
        array = [x.replace(char,"").lower() for x in array]
    for idx in range(len(array)):
        if "-" not in array[idx]:
            continue
        split = array[idx].split("-")
        array[idx] = split[0]
        for i in range(1,len(split)):
            array.insert(idx+i,split[i])

class Markov:
    def __init__(self,text=None,filename=None):
        """Creates a Markov object"""
        if verbose:
            print('creating markov object')
        # prepare the text for analysis
        if text:
            if verbose:
                print('creating text list')
            for char in "'\".,;_[]{}\\/+=|`~<>?!()*&^%$#@\"":
                text = text.replace(char,"").lower()
            text = text.replace("-"," ").lower()
            self.text_list = text.split()
        else:
            self.text_list = []
        self.word_freqs = {}
        self.word_vector = {}
        self.filename = filename

    def get_word_freqs(self):
        """Count the frequency of words in a text"""
        if verbose:
            print('getting word freqs')
        for word in self.text_list:
            if word in self.word_freqs:
                self.word_freqs[word] += 1
            else:
                self.word_freqs[word] = 1
        return self.word_freqs

    def print_word_freqs(self):
        for word in self.word_freqs:
            print(word,":", self.word_freqs[word])

    """def get_word_vector(self,num):
        if verbose:
            print('getting word vector')
        for idx in range(len(self.text_list)):
            if self.text_list[idx] in self.word_vector:
                print(self.text_list[idx], "in self.word_vector")
            else:
                self.word_vector[self.text_list[idx]] = {}
                print(self.text_list[idx],"now in self.word_vector")"""

    def get_word_vector(self,num):
        #Get word frequencies and collocations
        if verbose:
            print('getting word vector')
        for idx in range(len(self.text_list)):
            if self.text_list[idx] in self.word_vector:
                if idx + num > len(self.text_list):
                    end = len(self.text_list)
                else:
                    end = idx + num
                for n in range(idx+1,end):
                    if self.text_list[n] in self.word_vector[self.text_list[idx]][n-idx-1]:
                        self.word_vector[self.text_list[idx]][n-idx-1][self.text_list[n]] += 1
                    else:
                        self.word_vector[self.text_list[idx]][n-idx-1][self.text_list[n]] = 1
            else:
                self.word_vector[self.text_list[idx]] = []
                if idx + num > len(self.text_list):
                    end = len(self.text_list)
                else:
                    end = idx + num
                for r in range(num-1):
                    self.word_vector[self.text_list[idx]].append({})
                for n in range(idx+1,end):
                    self.word_vector[self.text_list[idx]][n-idx-1][self.text_list[n]] = 1
        return self.word_vector

    def return_word_vector(self):
        return self.word_vector

    def print_word_vector(self):
        for item in self.word_vector:
            print(item)
            for idx in range(len(self.word_vector[item])):
                print(idx, end=": ")
                for word in self.word_vector[item][idx]:
                    print(word,end=", ")
                print()

    def get_words(self,word,idx):
        """Show the most likely words to come after a given word"""
        if verbose:
            print('getting words')
        words = []
        for thing in self.word_vector[word][idx]:
            words.append([thing,self.word_vector[word][idx][thing]])
        words.sort(key=lambda x: x[1],reverse=True)
        return words

    def get_weight_freq(self):
        """Adjust the weights based on frequency"""
        if verbose:
            print('getting weight freq')
        for word in self.word_vector:
            freq = self.word_freqs[word]
            for array in self.word_vector[word]:  # first, second, third, etc. words
                for subitem in array:  # words that occur
                    array[subitem] = array[subitem]/freq
        return self.word_vector

    def change_weights(self,weight):
        """Change weights"""
        if verbose:
            print('changing weights')
        for word in self.word_vector:
            for item in self.word_vector[word]:  # first, second, third, etc. words
                for subitem in item:  # words that occur
                    item[subitem] = item[subitem] * weight
        return self.word_vector

    def output_sheet(self,filename):
        if verbose:
            print('getting output sheet')
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = 'Markov Weights'
        row = 0
        for word in self.word_vector:
            row += 1
            cell = ws1.cell(row=row,column=1)
            cell.value = word
            for col in range(2,len(self.word_vector[word])+1):
                cell = ws1.cell(row=row,column=col)
                value = ""
                for item in self.word_vector[word][col-1]:
                    value += item + ":" + str(self.word_vector[word][col-1][item]) + ","
                cell.value = value
        wb.save(filename=filename)

    def read_sheet(self):
        if verbose:
            print('reading sheet')
        wb = openpyxl.load_workbook(self.filename)
        ws1 = wb.active
        for row in range(1,len(ws1.rows)):
            word = ws1.cell(row=row,column=1).value
            """if not word.isalpha():
                continue"""
            print(word)
            self.word_vector[word] = []
            for column in range(2, len(ws1.columns)):
                self.word_vector[word].append({})
                text = ws1.cell(row=row,column=column).value
                entries = [x.split(":") for x in text.split(",")]
                for item in entries:
                    if item==['']:
                        continue
                    self.word_vector[word][column-2][item[0]] = item[1]

'''acd_file = open('acdsherlock.txt')
acd = ""
for line in acd_file:
    acd += str(line)
acd_Markov = Markov(acd)
acd_Markov.get_word_freqs()
print(len(acd_Markov.text_list))
print(len(acd_Markov.word_freqs))
acd_Markov.get_word_vector(5)
print(len(acd_Markov.return_word_vector()))
acd_Markov.get_weight_freq()
print(acd_Markov.get_words('the',0))
acd_Markov.output_sheet('acd_output.xlsx')
#acd_Markov.print_word_vector()'''
this_Markov = Markov(filename='/Users/cat/Downloads/acd_output-5.xlsx')
this_Markov.read_sheet()
print(this_Markov.get_words('the',0))
