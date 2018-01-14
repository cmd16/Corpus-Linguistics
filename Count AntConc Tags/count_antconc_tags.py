import openpyxl
import operator
import os

tag_list = ["CC", "CD", "DT", "EX", "FW", 'IN', 'IN/that', 'JJ', 'JJR', 'JJS', 'LS', 'MD', 'NN', 'NNS', 'NP', 'NPS', 'PDT', 'POS',
            'PP', 'PP$', 'RB', 'RBR', 'RBS', 'RP', 'SENT', 'SYM', 'TO', 'UH', 'VB', 'VBD', 'VBG', 'VBN', 'VBZ', 'VBP',
            'VD', 'VDD', 'VDG', 'VDN', 'VDZ', 'VDP', 'VH', 'VHD', 'VHG', 'VHN', 'VHZ', 'VHP', 'VV', 'VVD', 'VVG', 'VVN',
            'VVP', 'VVZ', 'WDT', 'WP', 'WP$', 'WRB', ':', '$', ',', '(', ')', "''", "``", "NONE"]

pos_definitions = {"CC": "coordinating conjuction", "CD":"cardinal number", "DT":"determiner", "EX":"existential there",
                   "FW":"foreign word", 'IN':"preposition/subord. conj", 'IN/that':"complementizer", 'JJ':"adjective",
                   'JJR': "adjective, comparative", 'JJS': "adjective, superlative", 'LS': "list marker", 'MD': "modal",
                   'NN': "noun, singular or mass", 'NNS': "noun, plural", 'NP': "proper noun, singular", 'NPS': "proper noun, plural",
                   'PDT':"predeterminer", 'POS':"possessive ending", 'PP':"personal pronoun", 'PP$':"possesive pronoun",
                   'RB': "adverb", 'RBR':"adverb, comparative", 'RBS':"adverb, superlative", 'RP':"particle", 'SENT':"end punctuation",
                   'SYM': "symbol", 'TO': "to", 'UH':"interjection", 'VB':"be", 'VBD':"was/were", 'VBG':"being", 'VBN':"been",
                   'VBZ':"is", 'VBP':"am/are", 'VD':"do", 'VDD':"did", 'VDG':"doing", 'VDN':"done", 'VDZ':"does", 'VDP':"do",
                   'VH':"have, base form", 'VHD':"had", 'VHG':"having", 'VHN':"had", 'VHZ':"has", 'VHP':"have, present non-3rd person",
                   'VV':"verb, base form", 'VVD':"verb, past tense", 'VVG': "verb, gerund/participle", 'VVN': "verb, past participle",
                   'VVP': "verb, present non-3rd person", 'VVZ': "verb, present 3rd person singular", 'WDT': "wh-determiner",
                   'WP': "wh-pronoun", 'WP$': "possessive wh-pronoun", 'WRB': "wh-adverb", ':' : "general joiner",
                   '$': "currency symbol", ',':",", '(':"(", ')':")", "''":"quotation marks", "NONE": "miscellaneous"}

test = True


# function definitions
def get_tag_dict(in_name):
    """
    Returns a dictionary that maps each part of speech (POS) to a dictionary containing words with that POS with their frequency,
    with a separate entry for each POS.
    :param in: the name of a .txt file created using TagAnt or a list of filenames of txt files that were created using TagAnt
    :return: a dictionary that maps each part of speech (POS) to a dictionary containing words with that POS with their frequency,
    with a separate entry for each POS
    """
    if type(in_name) == list:
        iterator = iter(in_name)
    else:
        iterator = iter([in_name])
    tag_dict = {tag: {} for tag in tag_list}
    for infilename in iterator:
        infile = open(infilename)
        for line in infile:
            items = line.split()  # get the words and tags by splitting on whitespace
            for item in items:
                word_pos = item.split("_")  # separate the word from the tag
                word = word_pos[0].lower()  # make the word lowercase, so "The" becomes "the"
                pos = word_pos[1]
                if pos == "``":
                    pos = "''"  # `` is equivalent to ''
                entry = tag_dict[pos]
                if word in entry:
                    entry[word] += 1
                else:
                    entry[word] = 1  # create the word
    return tag_dict


def print_tag_dict(aDict):
    """
    Prints out the abstract nouns in a dictionary, formatted nicely.
    :param aDict: a dictionary created by running get_tag_dict
    :return:
    """
    for pos in aDict:
        print(pos, "words:")
        for word in aDict[pos]:
            print(word, aDict[pos][word])  # print word and frequency
        print()  # new line for spacing


def store_spreadsheet(aDict, filename="Parts of Speech Spreadsheet", sort="alpha"):
    """
    Stores the instances of parts of speech and their frequencies in a spreadsheet.
    :param aDict: a dictionary created using get_tag_dict
    :param filename: the name of the spreadsheet to store the results in
    :param sort: indicates order the entries should be sorted in
    :return:
    """

    wb = openpyxl.Workbook()
    ws = wb.active  # get a handle to the sheet in the workbook

    ws['A1'] = "POS Spreadsheet"
    for col in range(1, len(tag_list) * 2 + 1, 2):  # go up by two because the suffixes need to be separated. add one because indexing starts at 1 in openpyxl
        pos = tag_list[(col - 1) // 2]
        entry = aDict[pos]
        if pos == "``":
            pos = "''"
        pos_def = pos_definitions[pos]
        ws.cell(row=2, column=col).value = pos_definitions[pos]
        # sort the dictionary
        if sort == "frequencyhi":
            entry_list = sorted(entry.items(), key=operator.itemgetter(1), reverse=True)
        elif sort=="alpha":
            entry_list = sorted(entry.items(), key=operator.itemgetter(0))
        elif sort=="reversealpha":
            entry_list = sorted(entry.items(), key=operator.itemgetter(0), reverse=True)
        elif sort=="frequencylo":
            entry_list = sorted(entry.items(), key=operator.itemgetter(1))
        for row in range(3, len(entry_list) + 3):  # this number is equal to the number of unique words with that suffix
            idx = row - 3
            word = entry_list[idx][0]
            ws.cell(row=row, column=col).value = word
            ws.cell(row=row, column=col+1).value = entry_list[idx][1]  # frequency
        if not filename.endswith(".xlsx"):
            wb.save(filename + ".xlsx")  # add the type extension if not included
        else:
            wb.save(filename)


def tag_dict_from_directory(path, walk=False):
    filename_list = []
    if walk:
        for item in os.walk(path):
            for filename in item[2]:
                if filename.endswith("_tagged.txt"):
                    filename_list.append(os.path.join(path, filename))
                    print(filename)
    else:
        for filename in os.listdir(path):
            if filename.endswith("_tagged.txt"):
                filename_list.append(os.path.join(path,filename))
    return get_tag_dict(filename_list)


if test:
    test_dict = tag_dict_from_directory("/Users/cat/Desktop/CalvinNewspapers",True)
    store_spreadsheet(test_dict, "test_multi_walk.xlsx", sort="frequencyhi")
