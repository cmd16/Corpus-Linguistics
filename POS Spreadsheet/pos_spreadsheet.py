import openpyxl
import operator
import os
from statistics import mean

tagant_tag_list = ["CC", "CD", "DT", "EX", "FW", 'IN', 'IN/that', 'JJ', 'JJR', 'JJS', 'LS', 'MD', 'NN', 'NNS', 'NP', 'NPS', 'PDT', 'POS',
            'PP', 'PP$', 'RB', 'RBR', 'RBS', 'RP', 'SENT', 'SYM', 'TO', 'UH', 'VB', 'VBD', 'VBG', 'VBN', 'VBZ', 'VBP',
            'VD', 'VDD', 'VDG', 'VDN', 'VDZ', 'VDP', 'VH', 'VHD', 'VHG', 'VHN', 'VHZ', 'VHP', 'VV', 'VVD', 'VVG', 'VVN',
            'VVP', 'VVZ', 'WDT', 'WP', 'WP$', 'WRB', ':', '$', ',', '(', ')', "''", "``", "NONE"]

tagant_pos_definitions = {"CC": "coordinating conjuction", "CD": "cardinal number", "DT": "determiner", "EX": "existential there",
                   "FW":"foreign word", 'IN':"preposition/subord. conj", 'IN/that':"complementizer", 'JJ':"adjective",
                   'JJR': "adjective, comparative", 'JJS': "adjective, superlative", 'LS': "list marker", 'MD': "modal",
                   'NN': "noun, singular or mass", 'NNS': "noun, plural", 'NP': "proper noun, singular", 'NPS': "proper noun, plural",
                   'PDT':"predeterminer", 'POS':"possessive ending", 'PP':"personal pronoun", 'PP$':"possesive pronoun",
                          'RB': "adverb", 'RBR':"adverb, comparative", 'RBS':"adverb, superlative", 'RP':"particle", 'SENT':"end punctuation",
                          'SYM': "symbol", 'TO': "to", 'UH':"interjection", 'VB':"be", 'VBD':"was/were", 'VBG':"being", 'VBN':"been",
                          'VBZ':"is", 'VBP':"am/are", 'VD':"do", 'VDD':"did", 'VDG':"doing", 'VDN':"done", 'VDZ':"does", 'VDP':"do",
                          'VH':"have, base form", 'VHD':"had", 'VHG':"having", 'VHN':"had", 'VHZ':"has", 'VHP':"have, present non-3rd person",
                          'VV':"verb, base form", 'VVD':"verb, past tense", 'VVG': "verb, gerund/participle", 'VVN': "verb, past participle",
                          'VVP': "verb, present non-3rd person", 'VVZ': "verb, present 3rd person singular", 'WDT': "wh-determiner",
                          'WP': "wh-pronoun", 'WP$': "possessive wh-pronoun", 'WRB': "wh-adverb", ':' : "general joiner",
                          '$': "currency symbol", ',':",", '(':"(", ')':")", "''":"quotation marks", "NONE": "miscellaneous"}

test = False


# function definitions
def get_tag_dict(in_name, taglist, delimiter, case_sensitive=False):
    """
    Returns a dictionary that maps each part of speech (POS) to a dictionary containing words with that POS with their frequency,
    with a separate entry for each POS.
    :param in_name: the name of a .txt file that contains tagged words or a list of filenames of txt files that contain tagged words
    :param taglist: a list containing the tags that the file(s) in in_name contain(s)
    :param delimiter: the character that marks a tag. In TagAnt, the delimiter is _
    :param case_sensitive: if true, "The" is a different word from "the"
    :return: a dictionary that maps each part of speech (POS) to a dictionary containing words with that POS with their frequency,
    with a separate entry for each POS.
    """
    if type(in_name) == list:
        iterator = iter(in_name)
    else:
        iterator = iter([in_name])
    tag_dict = {tag: {} for tag in taglist}
    for infilename in iterator:
        infile = open(infilename)
        print("processing " + infilename, end="... ")
        for line in infile:
            items = line.split()  # get the words and tags by splitting on whitespace
            for item in items:
                index = item.rfind(delimiter)  # find the last delimiter (the one before the tag)
                word = item[:index]
                if not case_sensitive:
                    word = word.lower()
                pos = item[index+1:]
                if pos == "``":
                    pos = "''"  # `` is equivalent to ''
                try:
                    entry = tag_dict[pos]
                except KeyError:
                    if word == "#":
                        pos = "SYM"
                    else:
                        pos = input("{" + word + "} " + line + " not in dictionary. Please enter a valid tag: ")
                    entry = tag_dict[pos]
                if word in entry:
                    entry[word] += 1
                else:
                    entry[word] = 1  # create the entry
        print("done")
    print("dictionary complete")
    return tag_dict


def count_sentences(in_name, sentence_marker):
    """
    Counts the number of sentences in a tagged text
    :param in_name: the name of a .txt file that contains tagged words or a list of filenames of txt files that contain tagged words
    Note: if a list of filenames is presented, this will count all the sentences across all the files. If you want individual
    counts for each file, call this function on each file.
    :param sentence_marker: the tag for the end of a sentence. In TagAnt, this is _SENT
    :return: the number of sentences
    """
    if type(in_name) == list:
        iterator = iter(in_name)
    else:
        iterator = iter([in_name])
    sentence_num = 0
    for infilename in iterator:
        infile = open(infilename)
        print("processing " + infilename)
        for line in infile:
            sentences = line.split(sentence_marker)
            for sent in sentences:
                words = sent.split()
                num_words = len(words)
                if num_words == 0:  # don't count blank lines
                    continue
                sentence_num += 1
    return sentence_num


def print_tag_dict(aDict):
    """
    Prints out the abstract nouns in a dictionary, formatted nicely.
    :param aDict: a dictionary created by running get_tag_dict
    :return:
    """
    for pos in aDict:
        print(pos, "words:")
        for word in aDict[pos]:
            print(word, aDict[pos][word])  # print word and frequency
        print()  # new line for spacing


def store_spreadsheet(aDict, taglist, tagdefs, filename="Parts of Speech Spreadsheet", sort="frequencyhi"):
    """
    Store the instances of various parts of speech and their frequencies in a spreadsheet
    :param aDict: a dictionary created using get_tag_dict
    :param taglist: a list containing the tags that the file(s) in in_name contain(s)
    :param tagdefs: a dictionary mapping the part of speech tags to their names/meanings
    :param filename: the name of the spreadsheet to store the results in
    :param sort: indicates order the entries should be sorted in
    :return:
    """

    wb = openpyxl.Workbook()
    ws = wb.active  # get a handle to the sheet in the workbook

    ws['A1'] = "POS Spreadsheet"
    for col in range(1, len(taglist) * 2 + 1, 2):  # go up by two because the POSs need to be separated. add one because indexing starts at 1 in openpyxl
        pos = taglist[(col - 1) // 2]
        print("processing", pos, "words")
        entry = aDict[pos]
        if pos == "``":
            pos = "''"
        pos_def = tagant_pos_definitions[pos]
        ws.cell(row=2, column=col).value = tagdefs[pos]
        # sort the dictionary
        if sort == "frequencyhi":
            entry_list = sorted(entry.items(), key=operator.itemgetter(1), reverse=True)
        elif sort=="alpha":
            entry_list = sorted(entry.items(), key=operator.itemgetter(0))
        elif sort=="reversealpha":
            entry_list = sorted(entry.items(), key=operator.itemgetter(0), reverse=True)
        elif sort=="frequencylo":
            entry_list = sorted(entry.items(), key=operator.itemgetter(1))
        for row in range(3, len(entry_list) + 3):  # this number is equal to the number of unique words with that suffix
            idx = row - 3
            word = entry_list[idx][0]
            try:
                ws.cell(row=row, column=col).value = word
            except openpyxl.utils.exceptions.IllegalCharacterError:  # try broad?
                word = input(word + " invalid. New value: ")
                ws.cell(row=row, column=col).value = word
            ws.cell(row=row, column=col+1).value = entry_list[idx][1]  # frequency
        if not filename.endswith(".xlsx"):
            wb.save(filename + ".xlsx")  # add the type extension if not included
        else:
            wb.save(filename)


def tag_dict_from_directory(path, delimiter, taglist, end="_tagged.txt", case_sensitive=False, walk=False):
    """
    Make a dictionary that maps each part of speech (POS) to a dictionary containing words with that POS with their frequency,
    with a separate entry for each POS.
    :param path: the path to the directory you want to look at. Note: this will aggregate data across all the files. If you want individual
    dictionaries for each file, call get_tag_dict on each file.
    :param delimiter: the character that marks a tag. In TagAnt, the delimiter is _
    :param taglist: a list containing the tags that the file(s) in in_name contain(s)
    :param end: the filename ending that signifies a tagged file.
    :param case_sensitive: if true, "The" is a different word from "the"
    :param walk: if True, the function "walks" the directory - it goes into subfolders
    :return: a dictionary that maps each part of speech (POS) to a dictionary containing words with that POS with their frequency,
    with a separate entry for each POS.
    """
    currentdir = os.getcwd()
    filename_list = []
    if walk:
        for item in os.walk(path):
            print(item)
            currentpath = item[0]
            os.chdir(currentpath)  # change to the directory you are looking at. useful for reading and writing files
            for filename in item[2]:
                if filename.endswith(end):
                    filename_list.append(os.path.join(currentpath, filename))
                    print(filename)
    else:
        for filename in os.listdir(path):
            if filename.endswith("_tagged.txt"):
                filename_list.append(os.path.join(path, filename))
    os.chdir(currentdir)
    return get_tag_dict(filename_list, taglist=taglist, case_sensitive=case_sensitive, delimiter=delimiter)


if test:
    originalnames = [x[:len(x)-4] + "_tagged.txt" for x in os.listdir("/Users/cat/Documents/Corpus Linguistics/Les Mis works En")[1:]]
    for name in originalnames:
        print(name)
    maindir = os.getcwd()
    os.chdir("./Tag Les Mis")
    newnames = [x for x in os.listdir(os.getcwd())[1:] if x in originalnames]
    dictionary = get_tag_dict(os.listdir(os.getcwd())[1:], taglist=tagant_tag_list, delimiter="_", case_sensitive=False)
    os.chdir(maindir)
    store_spreadsheet(dictionary, tagant_tag_list, tagant_pos_definitions, filename="Tagged Les Mis", sort="frequencyhi")
    """
    maindir = os.path.join(os.getcwd(), "Tag Hamilton")
    for directory in os.listdir(maindir):
        print()
        if 'Tag' in directory:
            print(directory)
            dictionary = tag_dict_from_directory(os.path.join(maindir, directory), taglist=tagant_tag_list, delimiter="_", walk=True)
            store_spreadsheet(dictionary, tagant_tag_list, tagant_pos_definitions, filename=directory, sort="frequencyhi")
    """
    # test_dict = tag_dict_from_directory("/Users/cat/Desktop/Tag Tolkien", case_sensitive=False, walk=True)
    # store_spreadsheet(test_dict, "POS_Tolkien.xlsx", sort="frequencyhi")
