import nltk
import treetaggerwrapper
import operator
import csv
import openpyxl
import os
import math
import re
import spacy
import time

# TODO: HAVE PAST VS PAST PARTICIPLE
#1) build a TreeTagger wrapper:
# tagger = treetaggerwrapper.TreeTagger(TAGLANG='en')
tree_taglist = ["CC", "CD", "DT", "EX", "FW", 'IN', 'IN/that', 'JJ', 'JJR', 'JJS', 'LS', 'MD', 'NN', 'NNS', 'NP', 'NPS', 'PDT', 'POS',
            'PP', 'PP$', 'RB', 'RBR', 'RBS', 'RP', 'SENT', 'SYM', 'TO', 'UH', 'VB', 'VBD', 'VBG', 'VBN', 'VBZ', 'VBP',
            'VH', 'VHD', 'VHG', 'VHN', 'VHZ', 'VHP', 'VV', 'VVD', 'VVG', 'VVN',
            'VVP', 'VVZ', 'WDT', 'WP', 'WP$', 'WRB', ':', '$', ',', '(', ')', "''", "NONE"]  # removed VD tags and "``"
tree_tagdefs = {"CC": "coordinating conjuction", "CD": "cardinal number", "DT": "determiner", "EX": "existential there",
                   "FW":"foreign word", 'IN':"preposition/subord. conj", 'IN/that':"complementizer", 'JJ':"adjective",
                   'JJR': "adjective, comparative", 'JJS': "adjective, superlative", 'LS': "list marker", 'MD': "modal",
                   'NN': "noun, singular or mass", 'NNS': "noun, plural", 'NP': "proper noun, singular", 'NPS': "proper noun, plural",
                'PDT':"predeterminer", 'POS':"possessive ending", 'PP':"personal pronoun", 'PP$':"possesive pronoun",
                'RB': "adverb", 'RBR':"adverb, comparative", 'RBS':"adverb, superlative", 'RP':"particle", 'SENT':"end punctuation",
                'SYM': "symbol", 'TO': "to", 'UH':"interjection", 'VB':"be", 'VBD':"was/were", 'VBG':"being", 'VBN':"been",
                'VBZ':"is", 'VBP':"am/are", 'VH':"have, base form", 'VHD':"had", 'VHG':"having", 'VHN':"had", 'VHZ':"has",
                'VHP':"have, present non-3rd person", # removed VD tags
                'VV':"verb, base form", 'VVD':"verb, past tense", 'VVG': "verb, gerund/participle", 'VVN': "verb, past participle",
                'VVP': "verb, present non-3rd person", 'VVZ': "verb, present 3rd person singular", 'WDT': "wh-determiner",
                'WP': "wh-pronoun", 'WP$': "possessive wh-pronoun", 'WRB': "wh-adverb", ':' : "general joiner",
                '$': "currency symbol", ',':",", '(':"(", ')':")", "''":"quotation marks", "NONE": "miscellaneous"}


# TODO: MODIFY THE CALL TO TAGFILE
def tag_directory(in_dir, out_dir, walk=False, restart="", end=""):
    """
    Tag the files in a directory specified by in_dir and store them in a directory specified by out_dir
    :param in_dir: the directory to read files from
    :param out_dir: the directory to write files to
    :param walk: if True, walk through the subdirectories
    :param restart: the first file to tag (skips all others)
    :return:
    """
    original_dir = os.getcwd()
    try:
        os.chdir(in_dir)
    except FileNotFoundError:
        os.mkdir(in_dir)
        os.chdir(in_dir)
    current = os.getcwd()
    namelist = []
    if restart:
        found = False
    else:
        found = True
    for name in os.listdir(current)[1:]:
        if name == restart:
            found = True
        if found:
            namelist.append(name)
    for filename in namelist:
        if filename == end:
            print(filename, end)
            break
        print("tag_directory processing", filename)
        try:
            tagger.tag_file_to(filename, os.path.join(out_dir, filename[:len(filename)-4] + "_tagged.txt"))
        except FileNotFoundError:
            os.mkdir(out_dir)
            tagger.tag_file_to(filename, os.path.join(out_dir, filename[:len(filename) - 4] + "_tagged.txt"))
        except treetaggerwrapper.TreeTaggerError:  # most common error seems to be html tags
            print("removing html tags in", filename, "due to TreeTaggerError")
            with open(filename) as file:
                with open("temp.txt", 'w') as tempfile:
                    line1 = next(file)
                    if "@page" not in line1:
                        tempfile.write(line1)
                    body = file.read()
                    body = re.sub("<.*?>", "", body)
                    tempfile.write(body)
            os.replace("temp.txt", filename)
            tagger.tag_file_to(filename, os.path.join(out_dir, filename[:len(filename)-4] + "_tagged.txt"))
            print("tagged file after removing html")
    os.chdir(original_dir)  # change back in case later methods need to be in this directory


def tag_lemma_from_tree(in_name, case_sensitive=False):
    """
    Returns a dictionary that maps each part of speech (POS) to a dictionary containing words with that POS with their frequency,
    with a separate entry for each POS.
    :param in_name: the name of a .txt file created using or a list of filenames of txt files that contain tagged words
    :param case_sensitive: if true, "The" is a different word from "the"
    :return: a tuple containing 0) a dictionary that maps each part of speech (POS) to a dictionary containing words with that POS with their frequency,
    with a separate entry for each POS 1) a dictionary that maps each lemma to its frequency
    """
    if type(in_name) == list:
        iterator = iter(in_name)
    else:
        iterator = iter([in_name])
    tag_dict = {tag: {} for tag in tree_taglist}
    lemma_dict = {}
    for infilename in iterator:
        infile = open(infilename)
        print("tag_lemma_from_tree processing " + infilename, end="... ")
        for line in infile:
            data = line.split()  # get a list containing word, tag, lemma
            word = data[0]
            if not case_sensitive:
                word = word.lower()
            try:
                pos = data[1]
                if pos == "``":
                    pos = "''"
                if case_sensitive:
                    lemma = data[2]
                else:
                    lemma = data[2].lower()
            except IndexError:
                if word.startswith("<"):
                    continue  # e.g., <repdns text="nosuey.Beta" />
                else:
                    print(line, data, sep=" : ")
                    word = input("enter word: ")
                    pos = input("enter pos: ")
                    lemma = input("enter lemma: ")
            try:
                entry = tag_dict[pos]
            except KeyError:
                if word == "#":
                    pos = "SYM"
                elif word.startswith("<rep"):  # skip replaced dns
                    continue
                else:
                    # pos = input("{" + word + "} " + line + " not in dictionary. Please enter a valid tag: ")
                    pos = "NONE"
                entry = tag_dict[pos]
            if word in entry:
                entry[word] += 1
            else:
                entry[word] = 1
            if lemma in lemma_dict:
                lemma_dict[lemma] += 1
            else:
                lemma_dict[lemma] = 1
        print("done")
    print("dictionaries complete")
    return tag_dict, lemma_dict


def tag_lemma_from_tree_directory(path, end="_tagged.txt", case_sensitive=False, walk=False):
    """
    Make a dictionary that maps each part of speech (POS) to a dictionary containing words with that POS with their frequency,
    with a separate entry for each POS.
    :param path: the path to the directory you want to look at. Note: this will aggregate data across all the files. If you want individual
    dictionaries for each file, call tag_lemma_from_tree on each file.
    :param end: the filename ending that signifies a tagged file.
    :param case_sensitive: if true, "The" is a different word from "the"
    :param walk: if True, the function "walks" the directory - it goes into subfolders
    :return: a tuple containing 0) a dictionary that maps each part of speech (POS) to a dictionary containing words with that POS with their frequency,
    with a separate entry for each POS 1) a dictionary that maps each lemma to its frequency
    """
    currentdir = os.getcwd()
    filename_list = []
    if walk:
        for item in os.walk(path):
            print(item)
            currentpath = item[0]
            os.chdir(currentpath)  # change to the directory you are looking at. useful for reading and writing files
            for filename in item[2]:
                if filename.endswith(end):
                    filename_list.append(os.path.join(currentpath, filename))
    else:
        for filename in os.listdir(path):
            if filename.endswith("_tagged.txt"):
                filename_list.append(os.path.join(path, filename))
    os.chdir(currentdir)
    return tag_lemma_from_tree(filename_list, case_sensitive=case_sensitive)


def store_taglemma_results(tag_lemma, filename="Parts of Speech Spreadsheet", sort="frequencyhi"):
    """
    Store the various parts of speech and their frequencies, as well as lemmas and their frequencies, in a spreadsheet
    :param tag_lemma: a tuple containing 0) a dictionary that maps each part of speech (POS) to a dictionary containing words with that POS with their frequency,
    with a separate entry for each POS 1) a dictionary that maps each lemma to its frequency [the return value of tag_lemma_from_tree]
    :param filename: the name of the spreadsheet to store the results in
    :param sort: indicates order the entries should be sorted in
    :return:
    """
    tag_dict = tag_lemma[0]
    lemma_dict = tag_lemma[1]
    wb = openpyxl.Workbook()
    ws = wb.active  # get a handle to the sheet in the workbook
    ws['A1'] = "POS and Lemma Spreadsheet"
    endcol = len(tree_taglist) * 2 + 1
    for col in range(1, endcol, 2):  # go up by two because the POSs need to be separated. add one because indexing starts at 1 in openpyxl
        pos = tree_taglist[(col - 1) // 2]
        print("processing", pos, "words")
        if pos == "``":
            pos = "''"  # `` is equivalent to ''
        entry = tag_dict[pos]
        ws.cell(row=2, column=col).value = tree_tagdefs[pos]
        # sort the dictionary
        if sort == "frequencyhi":
            entry_list = sorted(entry.items(), key=operator.itemgetter(1), reverse=True)
        elif sort == "alpha":
            entry_list = sorted(entry.items(), key=operator.itemgetter(0))
        elif sort == "reversealpha":
            entry_list = sorted(entry.items(), key=operator.itemgetter(0), reverse=True)
        elif sort == "frequencylo":
            entry_list = sorted(entry.items(), key=operator.itemgetter(1))
        for row in range(3, len(entry_list) + 3):  # this number is equal to the number of unique words with that suffix
            idx = row - 3
            word = entry_list[idx][0]
            try:
                ws.cell(row=row, column=col).value = word
            except openpyxl.utils.exceptions.IllegalCharacterError:
                # word = input(word + " invalid. New value: ")
                ws.cell(row=row, column=col).value = "IllegalCharacterError"
            ws.cell(row=row, column=col+1).value = entry_list[idx][1]  # frequency
        if not filename.endswith(".xlsx"):  # save the spreadsheet after each POS (in case of crash)
            wb.save(filename + ".xlsx")  # add the type extension if not included
        else:
            wb.save(filename)
    print("processing lemmas")
    ws.cell(row=2, column=endcol).value = "Lemmas"
    if sort == "frequencyhi":
        lemma_list = sorted(lemma_dict.items(), key=operator.itemgetter(1), reverse=True)
    elif sort == "alpha":
        lemma_list = sorted(lemma_dict.items(), key=operator.itemgetter(0))
    elif sort == "reversealpha":
        lemma_list = sorted(lemma_dict.items(), key=operator.itemgetter(0), reverse=True)
    elif sort == "frequencylo":
        lemma_list = sorted(lemma_dict.items(), key=operator.itemgetter(1))
    for row in range(3, len(lemma_list) + 3):
        idx = row - 3
        lemma = lemma_list[idx][0]
        try:
            ws.cell(row=row, column=endcol).value = lemma
        except openpyxl.utils.exceptions.IllegalCharacterError:
            # lemma = input(lemma + " invalid. New value: ")
            ws.cell(row=row, column=endcol).value = "IllegalCharacterError"
        ws.cell(row=row, column=endcol + 1).value = lemma_list[idx][1]  # frequency
    if not filename.endswith(".xlsx"):  # save the spreadsheet
        wb.save(filename + ".xlsx")  # add the type extension if not included
    else:
        wb.save(filename)


def freq_from_str(text, case_sensitive=False):
    """
    Get the frequencies of words from a string
    :param text: the text to count frequencies from
    :param case_sensitive: if False, convert everything to lowercase
    :return:
    """
    if case_sensitive:
        words = [word for word in nltk.word_tokenize(text) if
                 word.isalpha() or word.replace("'", "").isalpha()]
    else:
        words = [word for word in nltk.word_tokenize(text.lower()) if
                 word.isalpha() or word.replace("'", "").isalpha()]
    freqdist = nltk.FreqDist(words)
    return freqdist


def freq_from_txt(infile, case_senstive=False):
    """
    Get the frequencies of words from a txt file
    Limitations: I've noticed not all the contractions split correctly; e.g., "n't" and "shan't" appear in the same wordlist
    :param infile: the name of a txt file, or an open file handle
    :return: a FreqDist object with the unique words and their frequencies
    """
    try:
        f_in = open(infile)
        text = f_in.read()
        f_in.close()
    except TypeError:
        text = infile.read()
    except FileNotFoundError:
        print(infile)
        # print(infile[infile.index("all")+4:infile.index(".txt")])
        return False
    freqdist = freq_from_str(text, case_senstive)
    return freqdist


def freq_from_dir(path, case_sensitive=False, end=".txt", walk=False):
    """
    Get the frequencies of words from a directory. Data is aggreggated across all files.
    Limitations: if there are too many files, the program may run out of memory (all filenames are stored in a list).
    :param path: path to the directory to read
    :param case_sensitive: if True, "The" is different from "the"
    :param walk: if True, go into subdirectories
    :return: a FreqDist object with the unique words and their frequencies
    """
    currentdir = os.getcwd()
    filename_list = []
    if walk:
        for item in os.walk(path):
            # print(item)
            currentpath = item[0]  # might need to join with path
            os.chdir(currentpath)  # change to the directory you are looking at. useful for reading and writing files
            for filename in item[2]:
                if filename.endswith(end):
                    filename_list.append(os.path.join(currentpath, filename))
    else:
        for filename in os.listdir(path):
            if filename.endswith(end):
                filename_list.append(os.path.join(path, filename))
    os.chdir(currentdir)
    freqdist = nltk.FreqDist()
    for filename in filename_list:
        print("freq_from_dir getting freq from", filename)
        this_freqdist = freq_from_txt(filename, case_sensitive)
        if this_freqdist:  # if a freqdist was successfully created
            freqdist.update(this_freqdist)
    print()
    return freqdist


def freqdist_from_idfile (idfile, path, case_sensitive=False, end=".txt"):
    """
    Get the frequencies of words aggreggated across all files in an idlist file.
    :param idfile: the name of an idlist file (each line contains a relative filename without the extension).
    Can be an open file object or a filename.
    :param path: the directory to read the files from
    :param case_sensitive: if False, all words will be converted to lowercase
    :param end: the extension
    :return: a FreqDist object with the unique words and their frequencies
    """
    try:
        f_in = open(idfile, 'r')
        close = True  # tracks whether file should be closed
        filename = idfile
    except TypeError:
        f_in = idfile
        close = False
        filename = idfile.name
    freqdist = nltk.FreqDist()
    for line in f_in:
        name = os.path.join(path, line.strip() + end)
        this_freqdist = freq_from_txt(name, case_sensitive)
        if this_freqdist:  # will be false if the file couldn't open
            freqdist.update(this_freqdist)
        else:
            print(filename, line.strip())
    if close:
        f_in.close()
    return freqdist


def freqdist_to_wordlistfile(freqdist, filename):
    """
    Store the results of a frequency distribution in a txt file in a format similar to the one generated by AntConc

    :param freqdist: the FreqDist object
    :param filename: the txt file to store the results in. Must be a filename.
    :return:
    """
    tokens = freqdist.N()
    types = len(freqdist)
    # print(filename, tokens, types)
    with open(filename, 'w') as f_out:
        f_out.write("#Word types: " + str(types) + "\n")  # unique words
        f_out.write("#Word tokens: " + str(tokens) + "\n")  # total words
        f_out.write("#Search results: 0" + '\n')
        rank = 1
        for tup in freqdist.most_common():
            word = tup[0]
            freq = str(tup[1])
            f_out.write(str(rank) + "\t" + freq + "\t" + word + "\t\n")
            rank += 1


def wordlist_to_freqdist(wordlist_file):
    """
    Given a wordlist return a frequency distribution as well as the number of types and tokens
    :param wordlist_file: a .txt file generated from AntConc or freqdist_to_wordlistfile. Line 1 has the number of types,
    line 2 the number of tokens, and lines 4-end are in the format "rank\tword\tfrequency". Can be a filename or an open file object.
    :return: a frequency distribution (word frequencies), the number of types, and the number of tokens
    """
    try:
        corpus = open(wordlist_file)
        close = True
    except TypeError:
        corpus = wordlist_file
        close = False
    types = next(corpus)
    types = int(types[types.index("types")+7:])
    tokens = next(corpus)
    tokens = int(tokens[tokens.index("tokens")+8:])
    next(corpus)
    # load corpus into a FreqDist
    freqdist1 = nltk.FreqDist()
    for line in corpus:
        line = line.split()
        freqdist1[line[2]] = int(line[1])  # store the word and its frequency
    if close:
        corpus.close()
    return freqdist1


def freqdist_minus_freqdist(freqdist0, freqdist1):
    """
    Modify freqdist0 such that freqdist1 is removed from it
    :param freqdist0: FreqDist to modify
    :param freqdist1: FreqDist to compare
    :return:
    """
    raise NotImplementedError  # TODO: fix this
    for word in freqdist1:
        freqdist0[word] -= freqdist1[word]
        if freqdist0[word] == 0:
            freqdist0.remove(word)
    return freqdist0


def combine_wordlists_to_freqdist(wordlists):
    """
    Return a freqdist that is the combination of several wordlists
    :param wordlists: a list of wordlist filenames
    :return: Return a freqdist that is the combination of several wordlists
    """
    print("combining wordlists")
    freqdist = nltk.FreqDist()
    for wordlist in wordlists:
        this_freqdist = wordlist_to_freqdist(wordlist)
        freqdist.update(this_freqdist)
    return freqdist


def freqdist_to_excel(freqdist, filename):
    """
    Store the results of a frequency distribution in an excel spreadsheet
    :param freqdist: a frequency distribution with word frequencies, generated using
    :param filename: an excel filename to store the results in
    :return:
    """
    tokens = freqdist.N()
    types = len(freqdist)
    wb = openpyxl.Workbook()
    ws = wb.active  # get a handle to the sheet in the workbook
    ws['A1'] = "Freqdist Spreadsheet"
    ws['B1'] = "Total words (tokens)"
    ws['B2'] = tokens
    ws['B3'] = "Unique words (types)"
    ws['B4'] = types
    ws['C1'] = "word"
    ws['C2'] = "frequency"
    # ws['C3'] = "frequency per million"
    entries = freqdist.most_common()
    for rownum in range(3, len(entries) + 3):
        tup = entries[rownum - 3]  # a tuple with word, frequency, percentage (/10000 to get per million)
        ws.cell(row=rownum, col=1).value = tup[0]
        ws.cell(row=rownum, col=2).value = tup[1]
    wb.save(filename)
        # ws.cell(row=rownum, col=3).value = freqdist.freq[tup[0]] / 10000 # uncomment this for more data but it will go slower


def normalize_count_permillion(frequency, size):
    """
    Normalize a frequency given the frequency and the size
    :param frequency: frequency
    :param size: the size of the data (e.g., number of words in the corpus)
    :return: frequency per million
    """
    return frequency * 1000000 / size


def keyword_tuple_from_wordlists(corpus1name, corpus2name, p=0.01):
    """
    Keyword analysis from wordlist files.
    Limitations: the numbers are a little bit different here (about 10 lower or higher in the few examples I looked at)
    from the loglikelihood calculator at http://ucrel.lancs.ac.uk/cgi-bin/llsimple.pl?f1=3852&f2=2179&t1=3802120&t2=3569518
    even though I used the same formula. Probably due to different floating point arithmetic.
    Note: values are only stored if keyness is statistically significant as determined by the p value and
    normalized frequency 1 > normalized frequency 2
    :param corpus1name: name of the first txt wordlist file in the format generated by AntConc. Filename or open file object.
    :param corpus2name: name of the second txt wordlist file in the format generated by AntConc. Filename or open file object.
    :param p: p value (threshold for statistical significance)
    :return: a dictionary containing the words, keyness, raw frequencies and normalized frequencies.
    Now it is a tuple containing that dictionary and the types and tokens
    """
    print("keyword analysis from", corpus1name, "to", corpus2name)
    if p == 0.5:
        crit = 3.84
    elif p == 0.01:
        crit = 6.63
    elif p == 0.001:
        crit = 10.83
    elif p == 0.0001:
        crit = 15.13
    elif p == 0:
        crit = 0
    else:
        print(p, "not recognized as a valid p value (options are .5, .01, .001, .0001, and 0 for all results). Setting p value to .01")
        crit = 6.63  # the default is 1% error margin
    freqdist2 = wordlist_to_freqdist(corpus2name)
    types2, tokens2 = len(freqdist2), freqdist2.N()
    if tokens2 == 0:
        print(corpus2name, "is empty")
        return False # the wordlist is empty or something else went wrong
    keyword_dict = {}  # {word: (keyness, frequency1, normalizedfreq1, freq2, normalizedfreq2)}  # TODO: implement effect
    # start reading corpus1
    try:
        corpus1 = open(corpus1name)
        close = True
    except TypeError:
        corpus1 = corpus1name
        corpus1name = corpus1.name
        close = False
    types1 = int(next(corpus1)[13:])
    tokens1 = int(next(corpus1)[14:])
    if tokens1 == 0:
        print(corpus1name, "is empty")
        return False  # the wordlist is empty or something else went wrong
    next(corpus1)
    for line in corpus1:
        line = line.split()
        freq1 = int(line[1])
        word = line[2]
        freq2 = int(freqdist2[word])
        keyness, norm1, norm2 = calculate_keyness(freq1, freq2, tokens1, tokens2)
        if keyness >= crit and norm1 > norm2:
            keyword_dict[word] = (keyness, freq1, norm1, freq2, norm2)
    if close:
        corpus1.close()
    return keyword_dict, [corpus1name, types1, tokens1], [corpus2name, types2, tokens2]


def calculate_keyness(freq1, freq2, tokens1, tokens2):
    """
    Given both frequencies and the number of tokens, return the keyness as well as the normalized frequencies
    :param freq1: frequency of word in corpus1
    :param freq2: frequency of word in corpus2
    :param tokens1: number of tokens in corpus1
    :param tokens2: number of tokens in corpus2
    :return: keyness, normalized frequency 1, normalized frequency 2
    """
    num = (freq1 + freq2) / (tokens1 + tokens2)
    E1 = tokens1 * num
    E2 = tokens2 * num
    try:
        keyness = 2 * (freq1 * math.log(freq1 / E1) + (freq2 * math.log(freq2 / E2)))
    except ValueError:
        keyness = 2 * (freq1 * math.log(freq1 / E1))  # the second part equals 0
    norm1 = normalize_count_permillion(freq1, tokens1)
    norm2 = normalize_count_permillion(freq2, tokens2)
    return keyness, norm1, norm2


def keyword_tuple_from_freqdists(freqdist1, freqdist2, p=0.01, corpus1name="corpus 1", corpus2name="corpus 2"):
    """
    Keyword analysis from FreqDist objects.
    Limitations: the numbers are a little bit different here (about 10 lower or higher in the few examples I looked at)
    from the loglikelihood calculator at http://ucrel.lancs.ac.uk/cgi-bin/llsimple.pl?f1=3852&f2=2179&t1=3802120&t2=3569518
    even though I used the same formula. Probably due to different floating point arithmetic.
    Note: values are only stored if keyness is statistically significant as determined by the p value and
    normalized frequency 1 > normalized frequency 2
    :param freqdist1: FreqDist for corpus 1
    :param freqdist2: FreqDist for corpus 2
    :param p: p value (threshold for statistical significance)
    :param corpus1name: the name for corpus 1
    :param corpus2name: the name for corpus 2
    :return: a dictionary containing the words, keyness, raw frequencies and normalized frequencies.
    Now it is a tuple containing that dictionary and the types and tokens
    """
    if p == 0.5:
        crit = 3.84
    elif p == 0.01:
        crit = 6.63
    elif p == 0.001:
        crit = 10.83
    elif p == 0.0001:
        crit = 15.13
    elif p == 0:
        crit = 0
    else:
        print(p, "not recognized as a valid p value (options are .5, .01, .001, .0001, and 0 for all results). Setting p value to .01")
        crit = 6.63  # the default is 1% error margin
    tokens1 = freqdist1.N()
    if tokens1 == 0:
        print("freqdist1 is empty")
        return False
    tokens2 = freqdist2.N()
    if tokens2 == 0:
        print("freqdist2 is empty")
        return False
    types1 = len(freqdist1)
    types2 = len(freqdist2)
    keyword_dict = {}
    for word in freqdist1:
        freq1 = int(freqdist1[word])
        freq2 = int(freqdist2[word])
        keyness, norm1, norm2 = calculate_keyness(freq1, freq2, tokens1, tokens2)
        if keyness >= crit and norm1 > norm2:
            keyword_dict[word] = (keyness, freq1, norm1, freq2, norm2)
    return keyword_dict, [corpus1name, types1, tokens1], [corpus2name, types2, tokens2]


def keyword_tuple_from_keywordtxt(infile):
    """
    Read in a keywordtxt file (as created by store_keyword_txt) and store the results in a keyword tuple
    :param infile: a keywordtxt file (as created by store_keyword_txt). Filename or open file object.
    :return: a dictionary containing the words, keyness, raw frequencies and normalized frequencies.
    Now it is a tuple containing that dictionary and the types and tokens
    """
    keyword_dict = {}
    try:
        f_in = open(infile)
        close = True
    except TypeError:
        f_in = infile
        close = False
    except FileNotFoundError:
        raise FileNotFoundError
    for line in f_in:
        line = line.strip().split("\t")
        if line[0].startswith("#"):
            if line[0] == "# Corpus 1:":
                corpus1name, types1, tokens1 = line[1:]
            elif line[0] == "# Corpus 2:":
                corpus2name, types2, tokens2 = line[1:]
            continue
        word = line[0]
        keyness = float(line[1])
        try:
            freq1 = int(line[2])
        except ValueError:
            raise ValueError
        norm1 = float(line[3])
        freq2 = int(line[4])
        norm2 = float(line[5])
        keyword_dict[word] = (keyness, freq1, norm1, freq2, norm2)
    if close:
        f_in.close()
    return keyword_dict, [corpus1name, types1, tokens1], [corpus2name, types2, tokens2]


def store_keyword_txt(keyword_tuple, outfile, sort_key="keynesshi"):
    """
    Store a keyword dictionary in a txt file, separated by tabs. Note: values are only stored if keyness is statistically
    significant as determined by the p value and normalized frequency 1 > normalized frequency 2
    :param keyword_tuple: [keyword_dict, corpus1stats, corpus2stats]. The keyword dict contains keyness, raw frequencies,
    and normalized frequencies
    :param outfile: the name to store the keyword dictionary in. Filename or open file object.
    :param sort_key: the order to sort results in.
    :return:
    """
    try:
        f_out = open(outfile, 'w')
        close = True
    except TypeError:
        f_out = outfile
        close = False
    keyword_dict = keyword_tuple[0]
    corpus1name, types1, tokens1 = keyword_tuple[1]
    f_out.write("# Corpus 1:\t%s\t%d\t%d\n" % (corpus1name, types1, tokens1))
    corpus2name, types2, tokens2 = keyword_tuple[2]
    f_out.write("# Corpus 2:\t%s\t%d\t%d\n" % (corpus2name, types2, tokens2))
    f_out.write("# %s\t%s\t%s\t%s\t%s\t%s\n" % ("word", "keyness", "freq1", "norm1", "freq2", "norm2"))
    entry_list = []
    for item in keyword_dict.items():
        word = item[0]
        stats = item[1]
        keyness = stats[0]
        freq1 = stats[1]  # all of the stats numbers shift over 1 when you move the thing
        norm1 = stats[2]
        freq2 = stats[3]
        norm2 = stats[4]
        entry_list.append((word, keyness, freq1, norm1, freq2, norm2))
    if sort_key == "keynesshi":
        entry_list.sort(key=operator.itemgetter(1), reverse=True)
    elif sort_key == "keynesslo":
        entry_list.sort(key=operator.itemgetter(1))
    elif sort_key == "alphalo":
        entry_list.sort(key=operator.itemgetter(0))
    elif sort_key == "alphahi":
        entry_list.sort(key=operator.itemgetter(0), reverse=True)
    elif sort_key == "freq1hi":
        entry_list.sort(key=operator.itemgetter(2), reverse=True)
    elif sort_key == "freq1lo":
        entry_list.sort(key=operator.itemgetter(2))
    elif sort_key == "norm1hi":
        entry_list.sort(key=operator.itemgetter(3), reverse=True)
    elif sort_key == "norm1lo":
        entry_list.sort(key=operator.itemgetter(3))
    elif sort_key == "freq2hi":
        entry_list.sort(key=operator.itemgetter(4), reverse=True)
    elif sort_key == "freq2lo":
        entry_list.sort(key=operator.itemgetter(4))
    elif sort_key == "norm2hi":
        entry_list.sort(key=operator.itemgetter(5), reverse=True)
    elif sort_key == "norm2lo":
        entry_list.sort(key=operator.itemgetter(5))
    else:
        print("warning: sort key not understood (options are)", "setting sort key to keynesshi")  # TODO: print options
        entry_list.sort(key=operator.itemgetter(1), reverse=True)
    for entry in entry_list:
        word = entry[0]
        keyness = entry[1]
        freq1 = entry[2]  # all of the stats numbers shift over 1 when you move the thing
        norm1 = entry[3]
        freq2 = entry[4]
        norm2 = entry[5]
        f_out.write("%s\t%f\t%d\t%f\t%d\t%f\n" % (word, keyness, freq1, norm1, freq2, norm2))
    if close:
        f_out.close()


def find_similar_keywords(keyword_files, out_csv_name, sort_key="keyness1hi"):
    """
    Given a list of files of keyword dicts, find which keywords the corpora have in common
    master_key_dict is a dictionary that maps words to a list. mastery_key_dict[word] =
    [statsforcorpus1, statsforcorpus2, etc.]
    :param keyword_files: a list of filenames containing keyword analysis
    :param out_csv_name: the CSV to store results in
    :param sort_key: the key to sort the results by
    :return:
    """
    if len(keyword_files) < 2:
        raise ValueError("Must input 2 or more files for analysis")
    try:
        key_dict, [corpus1name, types1, tokens1], [corpus2name, types2, tokens2] = keyword_tuple_from_keywordtxt(keyword_files[0])
    except FileNotFoundError:
        print("FileNotFoundError", out_csv_name, keyword_files[0])
        return
    except ValueError:
        print("ValueError", out_csv_name, keyword_files[0])
        return
    master_key_dict = {}
    for word in key_dict:
        master_key_dict[word] = [key_dict[word]]
    for filename in keyword_files[1:]:
        try:
            key_dict, [corpus1name, types1, tokens1], [corpus2name, types2, tokens2] = keyword_tuple_from_keywordtxt(filename)
        except FileNotFoundError:
            print("FileNotFoundError", out_csv_name, filename)
            return
        except ValueError:
            print("ValueError", out_csv_name, filename)
            return
        words_to_del = []
        for word in master_key_dict:
            if word in key_dict:
                master_key_dict[word].append(key_dict[word])
            else:  # this isn't a keyword that is in all corpora, so we need to remove it
                words_to_del.append(word)
        for word in words_to_del:
            del master_key_dict[word]
    f_out = open(out_csv_name, 'w')
    writer = csv.writer(f_out, delimiter="\t", quotechar='|')
    filename_fields = []
    for filename in keyword_files:
        filename_fields.append(filename)
        filename_fields.extend([""]*4)
    filename_fields.pop()  # get rid of the last empty field
    writer.writerow([""] + filename_fields)  # first two are empty for word and keyness
    rows = []  # note: this is not memory efficient, but it's the simplest way to deal with the sorting problem
    for item in master_key_dict.items():
        word, stat_lists = item
        stat_flat = [word]
        for stat_list in stat_lists:
            keyness, freq1, norm1, freq2, norm2 = stat_list
            # ["%f" % keyness, "%d" % freq1, "%f" % norm1, "%d" % freq2, "%f" % norm2]
            stat_flat.extend(stat_list)
        rows.append(stat_flat)
    if sort_key == "alphalo":
        rows.sort(key=operator.itemgetter(0))
    elif sort_key == "keyness1hi":
        rows.sort(key=operator.itemgetter(1), reverse=True)
    elif sort_key == "alphahi":
        rows.sort(key=operator.itemgetter(0), reverse=True)
    elif sort_key == "keyness1lo":
        rows.sort(key=operator.itemgetter(1))
    else:
        rows.sort(key=operator.itemgetter(1), reverse=True)  # default to sorting by keynesshi
    for row in rows:
        str_row = []
        str_row.append(row.pop(0))  # append the word
        for i in range(0, len(keyword_files)*5, 5):
            keyness, freq1, norm1, freq2, norm2 = row[i:i+5]
            str_row.extend(["%f" % keyness, "%d" % freq1, "%f" % norm1, "%d" % freq2, "%f" % norm2])
        writer.writerow(str_row)
    f_out.close()


def find_prototypical(comp_freqdist, source_dir, idlist, outfilename, n=5, extension=False):
    """
    Find the n most prototypical files as compared to a FreqDist by performing keyword analysis and store those filenames
    :param freqdist: the FreqDist to compare
    :param source_dir: the directory to read files from
    :param idlist: a file with ids (relative filenames) to read. Filename or open file object.
    :param outfilename: the name of the file to store results in. Must be filename.
    :param n: the number of results to store
    :param extension: if False, omit the extension (e.g., .txt) when writing to outfilename
    and assume the extension has been omitted when reading from filenamelist
    :return: the full list of work_ids and their keyness
    """
    file_data = []
    try:
        f_in = open(idlist)
        close = True
    except TypeError:
        f_in = idlist
        close = False
    for line in f_in:
        id = line.strip()
        if extension:
            filename = os.path.join(source_dir, id)
        else:
            filename = os.path.join(source_dir, "%s.txt" % id)
        freqdist = freq_from_txt(filename)
        sum = 0
        keyword_dict0, [corpus1name, types1, tokens1], [corpus2name, types2, tokens2] = keyword_tuple_from_freqdists(freqdist,
            comp_freqdist, p=0)
        for word in keyword_dict0:
            keyness = keyword_dict0[word][0]
            sum += keyness
        keyword_dict1 = keyword_tuple_from_freqdists(comp_freqdist, freqdist, p=0)[0]
        for word in keyword_dict1:
            keyness = keyword_dict1[word][0]
            sum += keyness
            file_data.append((id, sum))
        if close:
            f_in.close()
    file_data.sort(key=operator.itemgetter(1), reverse=True)  # sort by high sum of keyword analysis
    out_file = open(outfilename, "w")
    for i in range(n):
        out_file.write("%s\t%f" % (file_data[i][0], file_data[i][1]))
    out_file.close()
    return file_data


def store_keyword_spreadsheet(keyword_tuple, filename="Keyword Spreadsheet.xlsx", sort="keyness1hi"):
    """
    Store a keyword dict to a spreadsheet
    :param keyword_tuple: keyword dict { word: (frequency, keyness)}, [stats1], [stats2]
    :param filename: the name of the spreadsheet
    :return:
    """
    print("storing keyword analysis")
    keyword_dict = keyword_tuple[0]
    corpus1_stats = keyword_tuple[1]
    corpus2_stats = keyword_tuple[2]
    wb = openpyxl.Workbook()
    ws = wb.active  # get a handle to the sheet in the workbook
    # ws['A1'] = "Keyword Spreadsheet"
    for idx in range(3):
        ws.cell(row=1, column=idx + 1).value = corpus1_stats[idx]
        ws.cell(row=1, column= idx + 4).value = corpus2_stats[idx]
    ws.cell(row=2, column=1).value = "Word"
    ws.cell(row=2, column=2).value = "Frequency 1"
    ws.cell(row=2, column=3).value = "Normalized frequency 1"
    ws.cell(row=2, column=4).value = "Frequency 2"
    ws.cell(row=2, column=5).value = "Normalized frequency 2"
    ws.cell(row=2, column=6).value = "Keyness"  # fixed error
    entry_list = []
    for item in keyword_dict.items():
        word = item[0]
        stats = item[1]
        freq1 = stats[0]  # all of the stats numbers shift over 1 when you move the thing
        norm1 = stats[1]
        freq2 = stats[2]
        norm2 = stats[3]
        keyness = stats[4]
        entry_list.append((word, freq1, norm1, freq2, norm2, keyness))  # frequency1, keyness
    if sort == "keynesshi":
        entry_list.sort(key=operator.itemgetter(5), reverse=True)
    else:  # TODO implement other sorting later
        entry_list.sort(key=operator.itemgetter(5), reverse=True)
    for row in range(3, len(entry_list) + 3):
        idx = row - 3
        word = entry_list[idx][0]
        freq1 = entry_list[idx][1]  # all of the stats numbers shift over 1 when you move the thing
        norm1 = entry_list[idx][2]
        freq2 = entry_list[idx][3]
        norm2 = entry_list[idx][4]
        keyness = entry_list[idx][5]
        try:
            ws.cell(row=row, column=1).value = word
        except openpyxl.utils.exceptions.IllegalCharacterError:
            word = input(word + " invalid. New value: ")
            ws.cell(row=row, column=1).value = word
        ws.cell(row=row, column=2).value = freq1  # store frequency
        ws.cell(row=row, column=3).value = norm1
        ws.cell(row=row, column=4).value = freq2
        ws.cell(row=row, column=5).value = norm2
        ws.cell(row=row, column=6).value = keyness
    if not filename.endswith(".xlsx"):
        wb.save(filename + ".xlsx")  # add the type extension if not included
    else:
        wb.save(filename)


def sent_with_word_file(proj_dir, word, dest_dir, filename, case_sensitive=False):
    """
    For a file in the fanfic directory, store all the sentences that include the specified word in a file in dest_dir
    :param proj_dir: the project directory (assumes the files to read are in os.path.join(proj_dir, "Fanfic_all"))
    :param word: the word or token to search for
    :param dest_dir: the directory to store the results files in
    :param filename: the file to read in. Must be a filename.
    :param case_sensitive: if False, all words will be converted to lowercase.
    :return:
    """
    f_in = open(os.path.join(proj_dir, "Fanfic_all/" + filename))
    text = f_in.read()
    f_in.close()
    if not case_sensitive:
        text = text.lower()
    f_out = open(os.path.join(dest_dir, filename), "w")
    sentences = nltk.tokenize.sent_tokenize(text)
    for sentence in sentences:
        if word in sentence and word in nltk.word_tokenize(sentence):
            f_out.write(sentence + "\n")
    f_out.close()


def sent_with_word_dir(proj_dir, word, dest_dir, case_sensitive=False, restart=None):
    """
    For each file in the fanfic directory, store all the sentences that include the specified word in a file in dest_dir
    :param proj_dir: the project directory (assumes the files to read are in os.path.join(proj_dir, "Fanfic_all"))
    :param word: the word or token to search for
    :param dest_dir: the directory to store the results files in
    :param case_sensitive: if False, all words will be converted to lowercase.
    :return:
    """
    fanfic_dir = os.path.join(proj_dir, "Fanfic_all")
    if restart:
        found = False
    for filename in os.listdir(fanfic_dir):
        if filename.endswith(".txt"):
            if restart and not found:
                if filename == restart:
                    found = True
                else:
                    print("skipping", filename)
                    continue
            print(filename)
            sent_with_word_file(proj_dir, word, dest_dir, filename, case_sensitive)


def head_of_word_to_freqdist(data_dir, word, nlp, dep=None, idfile=None):
    """
    Find the heads of a word by reading files created using sent_with_word_dir
    :param data_dir: the directory containing sentence files
    :param word: the target word
    :param nlp: the spacy object
    :param dep: the arc label, which describes the type of syntactic relation that connects the child to the head.
    This can be a list or a string
    :param idfile: an idlist file (so that only some files are read in). Filename or open file.
    :return: a FreqDist consisting of word heads and their frequencies
    """
    nlp = spacy.load("en")
    heads = []
    ids = []
    if idfile:
        try:
            f_in = open(idfile)
            lines = f_in.readlines()
            f_in.close()
        except TypeError:
            lines = idfile.readlines()
        ids.extend(lines)
    for filename in os.listdir(data_dir):
        if filename.endswith(".txt") and (not ids or filename in ids):
            print(filename)
            f_in = open(os.path.join(data_dir, filename))
            for sentence in f_in:
                doc = nlp(u'%s' % sentence)
                for token in doc:
                    if token.text == word:
                        if not dep or token.dep_ in dep:
                            head = token.head  # TODO: include tag/pos?
                            heads.append(head.text)
            f_in.close()
    freqdist = nltk.FreqDist(heads)
    return freqdist


def children_of_word_to_freqdist(data_dir, word, nlp, dep=None, idfile=None):
    """
    Find the children of a word by reading files created using sent_with_word_dir
    :param data_dir: the directory containing sentence files
    :param word: the target word
    :param nlp: the spacy object
    :param dep: the arc label, which describes the type of syntactic relation that connects the child to the head
    :param idfile: an idlist file (so that only some files are read in). Filename or open file.
    :return: a FreqDist consisting of word children and their frequencies
    """
    children = []
    ids = []
    if idfile:
        if idfile:
            try:
                f_in = open(idfile)
                lines = f_in.readlines()
                f_in.close()
            except TypeError:
                lines = idfile.readlines()
            ids.extend(lines)
    for filename in os.listdir(data_dir):
        if filename.endswith(".txt") and (not ids or filename in ids):
            print(filename)
            f_in = open(os.path.join(data_dir, filename))
            for sentence in f_in:
                doc = nlp(u'%s' % sentence)
                for token in doc:
                    if token.text == word:
                        if not dep or token.dep_ == dep:
                            kids = token.children  # TODO: include tag/pos?
                            children.extend(kids)
            f_in.close()
    freqdist = nltk.FreqDist(children)
    return freqdist


def pronoun_sents(proj_dir, pronouns, restart=None):
    """
    Get the sentences that contain the pronouns specified
    :param proj_dir: the project directory
    :param pronouns: the pronouns to search for
    :return:
    """
    for pronoun in pronouns:
        sent_with_word_dir(proj_dir, pronoun, os.path.join(proj_dir, "Pronoun sents/%s" % pronoun), restart=restart)


def pronoun_heads(proj_dir, nlp, pronouns, dep=None, idfile=None):
    """
    Make a FreqDist of the heads (grammatical dependency) of several pronouns
    :param proj_dir: the project directory
    :param nlp: the spacy object
    :param dep: the dependency to look at
    :param idfile: an idlist file (so that only some files are read in). Filename or open file.
    :return:
    """
    for pronoun in pronouns:
        print(pronoun)
        if pronoun in ["her", "woman", "man", "girl", "boy"]:  # could be possesive or direct object
            if pronoun != "her":  # "her" is not a subject
                freqdist = head_of_word_to_freqdist(os.path.join(proj_dir, "Pronoun sents/%s subj" % pronoun), nlp,
                                                    ["nsubj", "nsubjpass"], idfile)
                freqdist_to_wordlistfile(freqdist,
                                         os.path.join(proj_dir, "Pronoun data/%s_%s_python.txt" % (idfile, pronoun)))
            freqdist = head_of_word_to_freqdist(os.path.join(proj_dir, "Pronoun sents/%s obj" % pronoun), nlp,
                                                ["obj", "dobj", "iobj", "pobj"], idfile)
            freqdist_to_wordlistfile(freqdist,
                                     os.path.join(proj_dir, "Pronoun data/%s_%s_python.txt" % (idfile, pronoun)))
            freqdist = head_of_word_to_freqdist(os.path.join(proj_dir, "Pronoun sents/%s poss" % pronoun), nlp,
                                                ["poss"], idfile)
            freqdist_to_wordlistfile(freqdist,
                                     os.path.join(proj_dir, "Pronoun data/%s_%s_python.txt" % (idfile, pronoun)))
        else:
            freqdist = head_of_word_to_freqdist(os.path.join(proj_dir, "Pronoun sents/%s" % pronoun), nlp, dep, idfile)
            freqdist_to_wordlistfile(freqdist, os.path.join(proj_dir, "Pronoun data/%s_%s_python.txt" % (idfile, pronoun)))


def people_sents(proj_dir, csv_in, case_sensitive=False):
    fanfic_dir = os.path.join(proj_dir, "Fanfic_all")
    f_in = open(csv_in)
    reader = csv.reader(f_in)
    header = next(reader)
    stat_names = ['work_id', 'title', 'rating', 'category', 'fandom', 'relationship', 'character', 'additional tags',
                  'language',
                  'published', 'status', 'status date', 'words', 'chapters', 'comments', 'kudos', 'bookmarks',
                  'hits']  # no column for body
    header = {stat: stat_names.index(stat) for stat in stat_names}
    seen = []
    for row in reader:
        work_id = row[header["work_id"]]
        if work_id in seen:
            continue
        characters = row[header["character"]].split(", ")
    for filename in os.listdir(fanfic_dir):
        if filename.endswith(".txt"):
            print(filename)
            f_in = open(os.path.join(proj_dir, "Fanfic_all/" + filename))
            text = f_in.read()
            if not case_sensitive:
                text = text.lower()
            f_in.close()
    f_in.close()


if __name__ == "__main__":
    proj_dir = "/Volumes/2TB/Final_Project"

    # nlp = spacy.load("en")
    # print(nlp)
    # pronouns = ["she", "her", "he", "him", "his", "woman", "man", "girl", "boy"]
    # # pronouns_s = ["her", "he", "him", "his", "woman", "man", "girl", "boy"]
    # # pronoun_sents(proj_dir, ["boy"], restart="6992101.txt")
    # pronoun_heads(proj_dir, nlp, pronouns)
    # for idlist in os.listdir(os.path.join(proj_dir, "Fanfic lists")):
    #     if idlist.endswith(".txt"):
    #         pronoun_heads(proj_dir, nlp, pronouns, idfile=os.path.join(proj_dir, "Fanfic lists/" + idlist))
    # # print("BBC SH")
    # # freqdist = freqdist_from_idfile(os.path.join(proj_dir, "Fanfic lists/BBC Sherlock.txt"), os.path.join(proj_dir, "Fanfic_all"))
    # # freqdist_to_wordlistfile(freqdist, os.path.join(proj_dir, "wordlists/BBC_Sherlock_python.txt"))
